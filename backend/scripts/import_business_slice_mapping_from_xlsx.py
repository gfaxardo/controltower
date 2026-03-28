#!/usr/bin/env python3
"""
Importa ops.business_slice_mapping_rules desde Plantillas_Control_Tower_Simplificadas_final.xlsx
(hoja 1_Config_Tajadas). Trazabilidad: source_file_name, source_row_number.

Uso:
  cd backend && python -m scripts.import_business_slice_mapping_from_xlsx [--xlsx path] [--replace]
"""
from __future__ import annotations

import argparse
import logging
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from psycopg2.extras import execute_batch

from app.db.connection import get_db, init_db_pool

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger("import_business_slice")

HEADER_MARKERS = ("pais", "país", "ciudad", "tajada")


def _norm_yes(cell) -> bool:
    if cell is None:
        return False
    s = str(cell).strip().lower()
    return s in ("si", "sí", "yes", "true", "1")


def _split_csv(cell) -> list[str]:
    if cell is None:
        return []
    s = str(cell).strip()
    if not s or s.upper() == "INACTIVO":
        return []
    return [p.strip() for p in s.split(",") if p.strip() and p.strip().upper() != "INACTIVO"]


def _parse_fleet_tokens(cell) -> list[dict]:
    """Devuelve lista {fleet_display_name, is_subfleet, subfleet_name, parent_fleet_name}.

    Jerarquía: la primera flota principal (no sf) del listado es padre de todas las subflotas (sf).
    """
    if cell is None:
        return []
    raw = str(cell).strip()
    if not raw or raw.upper() == "INACTIVO":
        return []
    parts = re.split(r",(?![^(]*\))", raw)
    staged: list[tuple[str, bool]] = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        m = re.match(r"^(.+?)\s*\(\s*sf\s*\)\s*$", p, re.IGNORECASE)
        if m:
            staged.append((m.group(1).strip(), True))
        else:
            staged.append((p, False))
    first_main = next((name for name, is_sf in staged if not is_sf), None)
    out: list[dict] = []
    for name, is_sf in staged:
        if is_sf:
            out.append(
                {
                    "fleet_display_name": name,
                    "is_subfleet": True,
                    "subfleet_name": name,
                    "parent_fleet_name": first_main,
                }
            )
        else:
            out.append(
                {
                    "fleet_display_name": name,
                    "is_subfleet": False,
                    "subfleet_name": None,
                    "parent_fleet_name": None,
                }
            )
    return out or [
        {
            "fleet_display_name": raw,
            "is_subfleet": False,
            "subfleet_name": None,
            "parent_fleet_name": None,
        }
    ]


def _parse_parks(cell) -> list[str]:
    if cell is None:
        return []
    s = str(cell).strip()
    if not s or "INACTIVO" in s.upper():
        return []
    return [p.strip() for p in s.split(",") if p.strip() and "INACTIVO" not in p.upper()]


def _infer_rule_type(req_tipo: bool, req_works: bool, tipo_vals: list[str], works_vals: list[str]) -> str:
    if req_works and works_vals:
        return "park_plus_works_terms"
    if req_tipo and tipo_vals:
        return "park_plus_tipo_servicio"
    return "park_only"


def _find_header_row(ws) -> int:
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or row[0] is None:
            continue
        h0 = str(row[0]).strip().lower()
        if h0 in HEADER_MARKERS or "tajada" in h0:
            return i
    raise ValueError("No se encontró fila de encabezados (Pais/Ciudad/Tajada).")


def load_rows(xlsx_path: str) -> tuple[list[dict], str]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "1_Config_Tajadas" not in wb.sheetnames:
        raise ValueError(f"Hojas: {wb.sheetnames}; se esperaba 1_Config_Tajadas")
    ws = wb["1_Config_Tajadas"]
    hdr = _find_header_row(ws)
    out: list[dict] = []
    for ridx, row in enumerate(ws.iter_rows(min_row=hdr + 1, values_only=True), start=hdr + 1):
        if not row or row[0] is None:
            continue
        country = str(row[0]).strip() if row[0] else ""
        city = str(row[1]).strip() if row[1] else ""
        slice_name = str(row[2]).strip() if row[2] else ""
        if not slice_name or slice_name.upper() == "INACTIVO":
            continue
        fleet_cell = row[3] if len(row) > 3 else None
        parks_cell = row[4] if len(row) > 4 else None
        req_tipo = _norm_yes(row[5]) if len(row) > 5 else False
        tipo_vals = _split_csv(row[6]) if len(row) > 6 else []
        req_works = _norm_yes(row[7]) if len(row) > 7 else False
        works_raw = row[8] if len(row) > 8 else None
        works_vals = _split_csv(works_raw)
        notes = str(row[9]).strip() if len(row) > 9 and row[9] else None

        parks = _parse_parks(parks_cell)
        fleets = _parse_fleet_tokens(fleet_cell)
        row_inactive = (
            not parks
            or (fleet_cell and str(fleet_cell).strip().upper() == "INACTIVO")
            or (parks_cell and str(parks_cell).strip().upper() == "INACTIVO")
        )
        rule_type = _infer_rule_type(req_tipo, req_works, tipo_vals, works_vals)
        if rule_type == "park_plus_works_terms" and not works_vals:
            logger.warning(
                "Fila %s: requiere works_terms pero lista vacía (%s / %s). Se marca inactiva.",
                ridx,
                slice_name,
                city,
            )
            row_inactive = True
        if rule_type == "park_plus_tipo_servicio" and not tipo_vals:
            logger.warning(
                "Fila %s: requiere tipo_servicio pero lista vacía (%s / %s). Se marca inactiva.",
                ridx,
                slice_name,
                city,
            )
            row_inactive = True

        if not fleets:
            fleets = [
                {
                    "fleet_display_name": slice_name,
                    "is_subfleet": False,
                    "subfleet_name": None,
                    "parent_fleet_name": None,
                }
            ]

        if not parks:
            out.append(
                {
                    "country": country,
                    "city": city,
                    "business_slice_name": slice_name,
                    "fleet_display_name": fleets[0]["fleet_display_name"],
                    "is_subfleet": fleets[0]["is_subfleet"],
                    "subfleet_name": fleets[0]["subfleet_name"],
                    "parent_fleet_name": fleets[0]["parent_fleet_name"],
                    "park_id": "__INACTIVE_NO_PARK__",
                    "rule_type": rule_type,
                    "tipo_servicio_values": tipo_vals,
                    "works_terms_values": works_vals,
                    "notes": notes,
                    "source_row_number": ridx,
                    "is_active": False,
                }
            )
            continue

        for park_id in parks:
            for fl in fleets:
                out.append(
                    {
                        "country": country,
                        "city": city,
                        "business_slice_name": slice_name,
                        "fleet_display_name": fl["fleet_display_name"],
                        "is_subfleet": fl["is_subfleet"],
                        "subfleet_name": fl["subfleet_name"],
                        "parent_fleet_name": fl["parent_fleet_name"],
                        "park_id": park_id,
                        "rule_type": rule_type,
                        "tipo_servicio_values": tipo_vals,
                        "works_terms_values": works_vals,
                        "notes": notes,
                        "source_row_number": ridx,
                        "is_active": not row_inactive,
                    }
                )
    return out, os.path.basename(xlsx_path)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--xlsx",
        default=os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "exports",
            "Plantillas_Control_Tower_Simplificadas_final.xlsx",
        ),
    )
    ap.add_argument("--replace", action="store_true", help="TRUNCATE ops.business_slice_mapping_rules antes de insertar")
    args = ap.parse_args()

    if not os.path.isfile(args.xlsx):
        logger.error("No existe el archivo: %s", args.xlsx)
        return 1

    rows, src_name = load_rows(args.xlsx)
    logger.info("Filas generadas para insert: %s (desde %s)", len(rows), src_name)

    init_db_pool()
    sql = """
        INSERT INTO ops.business_slice_mapping_rules (
            country, city, business_slice_name, fleet_display_name, is_subfleet,
            subfleet_name, parent_fleet_name, park_id, rule_type,
            tipo_servicio_values, works_terms_values, notes,
            source_file_name, source_row_number, is_active
        ) VALUES (
            %(country)s, %(city)s, %(business_slice_name)s, %(fleet_display_name)s, %(is_subfleet)s,
            %(subfleet_name)s, %(parent_fleet_name)s, %(park_id)s, %(rule_type)s,
            %(tipo_servicio_values)s, %(works_terms_values)s, %(notes)s,
            %(source_file_name)s, %(source_row_number)s, %(is_active)s
        )
    """
    with get_db() as conn:
        cur = conn.cursor()
        if args.replace:
            cur.execute("TRUNCATE ops.business_slice_mapping_rules RESTART IDENTITY CASCADE")
            logger.info("Tabla truncada (replace).")
        for r in rows:
            r["source_file_name"] = src_name
            r["tipo_servicio_values"] = r["tipo_servicio_values"] or []
            r["works_terms_values"] = r["works_terms_values"] or []
        execute_batch(cur, sql, rows, page_size=200)
        conn.commit()
        cur.close()

    logger.info(
        "Import OK. Recalcular agregados: python -m scripts.refresh_business_slice_mvs "
        "(p. ej. --month YYYY-MM o --backfill-from / --backfill-to)."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
