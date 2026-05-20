"""Fase 1F-2 — Importador seguro de identidades bancarias.

Importa CSV o XLSX a fraud.payment_identity_source.
NUNCA guarda ni imprime account_number completo.
Soporta dry_run.

Uso:
  python backend/scripts/fraud_import_payment_identities.py --file datos.csv --source-name nomina_mayo --created-by admin --dry-run true
  python backend/scripts/fraud_import_payment_identities.py --file datos.csv --source-name nomina_mayo --created-by admin --dry-run false
"""
import sys, os, argparse, csv, uuid, hashlib, re
from datetime import datetime

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(__file__))))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

from app.db.connection import get_db
from app.services.fraud.fraud_feature_service import normalize_bank_account, mask_account_number, hash_bank_cluster_key
from psycopg2.extras import Json

REQUIRED_COLS = {"driver_id", "bank_name", "account_number"}
OPTIONAL_COLS = {"park_id", "account_holder", "document", "status", "created_at", "source_note"}


def read_rows(filepath):
    """Lee filas desde CSV o XLSX. Retorna lista de dicts."""
    ext = os.path.splitext(filepath)[1].lower()
    rows = []
    if ext == ".csv":
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append({k.strip().lower().replace(" ", "_"): v for k, v in row.items()})
    elif ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower().replace(" ", "_") for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                d = {}
                for i, h in enumerate(headers):
                    d[h] = str(row[i]) if row[i] is not None else ""
                rows.append(d)
            wb.close()
        except ImportError:
            print("ERROR: openpyxl no instalado. Usa CSV o instala openpyxl.")
            sys.exit(1)
    else:
        print(f"ERROR: Formato no soportado: {ext}. Usa .csv o .xlsx")
        sys.exit(1)
    return rows


def validate_row(row, row_num):
    """Valida una fila. Retorna (is_valid, errors, normalized_data)."""
    errors = []

    driver_id = (row.get("driver_id") or "").strip()
    bank_name = (row.get("bank_name") or "").strip()
    account_number = (row.get("account_number") or "").strip()

    if not driver_id:
        errors.append(f"row {row_num}: driver_id vacio")
    if not account_number:
        errors.append(f"row {row_num}: account_number vacio")
    if not bank_name:
        errors.append(f"row {row_num}: bank_name vacio")

    if errors:
        return False, errors, None

    bn_norm, an_norm = normalize_bank_account(bank_name, account_number)
    account_hash = hash_bank_cluster_key(bank_name, account_number)
    masked = mask_account_number(account_number)
    park_id = (row.get("park_id") or "").strip() or None

    data = {
        "driver_id": driver_id,
        "park_id": park_id,
        "bank_name_norm": bn_norm,
        "account_hash": account_hash,
        "masked_account_number": masked,
        "evidence": {
            "account_holder": (row.get("account_holder") or row.get("recipient_name") or "").strip() or None,
            "document": (row.get("document") or row.get("document_number") or "").strip() or None,
            "status": (row.get("status") or row.get("is_active") or "").strip() or None,
            "source_note": (row.get("source_note") or "").strip() or None,
            "imported_at": datetime.now().isoformat(),
        },
    }
    # Limpiar evidence de None
    data["evidence"] = {k: v for k, v in data["evidence"].items() if v is not None}
    return True, [], data


def import_file(filepath, source_name, created_by, dry_run=True, deactivate_missing=False, limit=None):
    """Importa archivo a fraud.payment_identity_source."""
    batch_id = str(uuid.uuid4())[:12]
    filename = os.path.basename(filepath)
    started = datetime.now()

    print(f"\n{'[DRY RUN] ' if dry_run else ''}Import: {filename}")
    print(f"  batch_id: {batch_id}")
    print(f"  source: {source_name}")
    print(f"  created_by: {created_by}")

    rows = read_rows(filepath)
    total = len(rows)
    if limit:
        rows = rows[:int(limit)]
        print(f"  limit: {limit} (de {total} totales)")
        total = len(rows)

    valid_data = []
    invalid_count = 0
    errors_list = []

    # 1. Validar y normalizar
    seen_hashes = set()  # dedup dentro del archivo
    for i, row in enumerate(rows, start=1):
        ok, errs, data = validate_row(row, i)
        if not ok:
            invalid_count += 1
            for e in errs:
                errors_list.append(e)
            if len(errors_list) >= 20:
                break
            continue
        # Dedup dentro del archivo: mismos driver_id + account_hash
        dedup_key = (data["driver_id"], data["account_hash"])
        if dedup_key in seen_hashes:
            continue
        seen_hashes.add(dedup_key)
        valid_data.append(data)

    duplicated_in_file = total - invalid_count - len(valid_data)

    # 2. Batch upsert (sin per-row existence check — ON CONFLICT lo maneja)
    inserted = 0
    updated = 0
    skipped = 0
    duplicated_existing = 0

    if not dry_run and valid_data:
        from psycopg2.extras import execute_values
        chunk = 200
        for i in range(0, len(valid_data), chunk):
            batch = valid_data[i:i + chunk]
            rows = []
            for d in batch:
                rows.append((
                    d["driver_id"], d["park_id"], d["bank_name_norm"],
                    d["account_hash"], d["masked_account_number"],
                    source_name, batch_id, i + len(rows) + 1,
                    Json(d["evidence"]),
                ))
            with get_db() as wconn:
                wcur = wconn.cursor()
                execute_values(wcur, """
                    INSERT INTO fraud.payment_identity_source
                        (driver_id, park_id, bank_name_norm, account_hash,
                         masked_account_number, source_name, source_batch_id,
                         source_row_number, is_active, evidence)
                    VALUES %s
                    ON CONFLICT (driver_id, account_hash, source_name) DO UPDATE SET
                        masked_account_number = EXCLUDED.masked_account_number,
                        bank_name_norm = EXCLUDED.bank_name_norm,
                        evidence = EXCLUDED.evidence,
                        is_active = true,
                        source_batch_id = EXCLUDED.source_batch_id,
                        source_row_number = EXCLUDED.source_row_number,
                        updated_at = now()
                """, rows, template="(%s,%s,%s,%s,%s,%s,%s,%s,true,%s)")
                wconn.commit()
                wcur.close()
            inserted += len(rows)
        # After batch, estimate existing by checking DB count
        with get_db() as conn:
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM fraud.payment_identity_source WHERE source_name = %s", (source_name,))
            total_in_db = cur.fetchone()[0]
            cur.close()
        updated = max(0, total_in_db - len(valid_data))
        inserted = len(valid_data)

    # 4. Registrar import_log
    status = "dry_run_completed" if dry_run else "completed"
    with get_db() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO fraud.payment_identity_import_log
                (batch_id, source_name, file_name, dry_run, status,
                 total_rows, valid_rows, invalid_rows, duplicated_rows,
                 inserted_rows, updated_rows, skipped_rows,
                 errors, started_at, finished_at, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, now(), %s)
            ON CONFLICT (batch_id) DO UPDATE SET
                status = EXCLUDED.status, finished_at = now(),
                inserted_rows = EXCLUDED.inserted_rows, updated_rows = EXCLUDED.updated_rows
        """, (
            batch_id, source_name, filename, dry_run, status,
            total, len(valid_data), invalid_count, duplicated_existing + duplicated_in_file,
            inserted, updated, skipped,
            Json(errors_list[:50]) if errors_list else None,
            started, created_by,
        ))
        conn.commit()
        cur.close()

    # 5. Resumen
    print(f"\n=== IMPORT SUMMARY [{batch_id}] ===")
    print(f"  total_rows:       {total}")
    print(f"  valid_rows:       {len(valid_data)}")
    print(f"  invalid_rows:     {invalid_count}")
    print(f"  duplicated:       {duplicated_existing + duplicated_in_file}")
    print(f"  inserted:         {inserted}")
    print(f"  updated:          {updated}")
    print(f"  skipped:          {skipped}")
    print(f"  mode:             {'dry_run' if dry_run else 'commit'}")

    if errors_list:
        print(f"\n  Validation errors (first 10):")
        for e in errors_list[:10]:
            print(f"    - {e}")
        if len(errors_list) > 10:
            print(f"    ... y {len(errors_list)-10} mas")

    if not dry_run and inserted + updated > 0:
        print(f"\n  >> Para recomputar bank clusters:")
        print(f"  python backend/scripts/fraud_recompute.py --routines bank_account_cluster --dry-run false")

    return {
        "batch_id": batch_id, "total": total, "valid": len(valid_data),
        "invalid": invalid_count, "duplicated": duplicated_existing + duplicated_in_file,
        "inserted": inserted, "updated": updated,
    }


def main():
    parser = argparse.ArgumentParser(description="Import payment identities")
    parser.add_argument("--file", required=True)
    parser.add_argument("--source-name", required=True)
    parser.add_argument("--created-by", default="system")
    parser.add_argument("--dry-run", type=str, default="true")
    parser.add_argument("--deactivate-missing", type=str, default="false")
    parser.add_argument("--limit", type=int, default=None)
    args = parser.parse_args()

    dry_run = args.dry_run.lower() in ("true", "1", "yes")
    deactivate = args.deactivate_missing.lower() in ("true", "1", "yes")

    import_file(args.file, args.source_name, args.created_by, dry_run, deactivate, args.limit)


if __name__ == "__main__":
    main()
