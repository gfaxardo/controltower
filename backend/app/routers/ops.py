import threading
from contextlib import suppress

from fastapi import APIRouter, Query, HTTPException, Body, Request
import os
import json
import time as _time

def _debug_log_ops(location: str, message: str, data: dict, hypothesis_id: str = ""):
    try:
        log_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "..", "debug-9075f8.log")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(json.dumps({"sessionId": "9075f8", "location": location, "message": message, "data": data, "timestamp": int(_time.time() * 1000), "hypothesisId": hypothesis_id}) + "\n")
    except Exception:
        pass

from app.services.ops_universe_service import get_ops_universe
from app.services.territory_quality_service import (
    get_territory_kpis_total,
    get_territory_kpis_weekly,
    get_unmapped_parks
)
from app.services.plan_vs_real_service import (
    get_plan_vs_real_monthly,
    get_alerts_monthly,
    get_latest_parity_audit,
    log_plan_vs_real_source_usage,
)
from app.services.control_loop_plan_vs_real_service import (
    get_control_loop_plan_vs_real,
    list_control_loop_plan_versions,
)
from app.services.projection_expected_progress_service import (
    get_omniview_projection,
)
from app.services.business_slice_real_freshness_service import (
    get_omniview_business_slice_real_freshness,
)
from app.services.business_slice_real_refresh_job import run_business_slice_real_refresh_job
from app.utils.json_sanitizer import sanitize_for_json
from app.settings import settings
from app.services.plan_real_split_service import (
    get_real_monthly,
    get_plan_monthly,
    get_overlap_monthly
)
from app.services.real_lob_service import (
    get_real_lob_monthly as get_real_lob_monthly_svc,
    get_real_lob_weekly as get_real_lob_weekly_svc,
    get_real_lob_meta,
)
from app.services.real_lob_service_v2 import (
    get_real_lob_monthly_v2,
    get_real_lob_weekly_v2,
    get_real_lob_meta_v2,
)
from app.services.real_lob_filters_service import get_real_lob_filters
from app.services.real_lob_v2_data_service import get_real_lob_v2_data
from app.services.real_strategy_service import (
    get_real_strategy_country,
    get_real_strategy_lob,
    get_real_strategy_cities,
)
from app.services.real_drill_service import (
    get_real_drill_summary,
    get_real_drill_summary_countries,
    get_real_drill_by_lob,
    get_real_drill_by_park,
    get_real_drill_totals,
    get_real_drill_meta,
    get_real_drill_coverage,
    refresh_real_drill_mv,
    RealDrillMvNotPopulatedError,
)
from app.services.real_lob_drill_pro_service import (
    get_drill as get_real_lob_drill_pro,
    get_drill_children as get_real_lob_drill_pro_children,
    get_drill_parks as get_real_lob_drill_parks,
)
from app.services.period_semantics_service import get_period_semantics
from app.services.comparative_metrics_service import (
    get_weekly_comparative,
    get_monthly_comparative,
)
from app.services.real_lob_daily_service import (
    get_daily_summary,
    get_daily_comparative,
    get_daily_table,
)
from app.services.real_operational_service import (
    get_operational_snapshot,
    get_day_view,
    get_hourly_view,
    get_cancellation_view,
)
from app.services.real_operational_comparatives_service import (
    get_today_vs_yesterday,
    get_today_vs_same_weekday_avg,
    get_current_hour_vs_historical,
    get_this_week_vs_comparable,
)
from app.services.supply_service import (
    get_supply_geo,
    get_supply_parks,
    get_supply_series,
    get_supply_summary,
    get_supply_global_series,
    get_supply_segments_series,
    get_supply_segment_config,
    get_supply_alerts,
    get_supply_alert_drilldown,
    refresh_supply_alerting_mvs,
    get_supply_overview_enhanced,
    get_supply_composition,
    get_supply_migration,
    get_supply_migration_drilldown,
    get_supply_migration_weekly_summary,
    get_supply_migration_critical,
    get_supply_freshness,
)
from app.services.data_freshness_service import (
    get_freshness_audit,
    get_freshness_alerts,
    get_freshness_expectations,
    get_freshness_global_status,
)
from app.services.data_integrity_service import (
    get_integrity_report,
    get_system_health,
)
from app.services.real_margin_quality_service import get_margin_quality_full
from app.services.supply_definitions import get_definitions
from app.services.behavior_alerts_service import (
    get_behavior_alerts_summary,
    get_behavior_alerts_drivers,
    get_behavior_alerts_driver_detail,
    get_behavior_alerts_export,
    get_behavior_alerts_insight,
)
from app.services.action_engine_service import (
    get_action_engine_summary,
    get_action_engine_cohorts,
    get_action_engine_cohort_detail,
    get_action_engine_recommendations,
    get_action_engine_export,
)
from app.services.driver_behavior_service import (
    get_driver_behavior_summary,
    get_driver_behavior_drivers,
    get_driver_behavior_driver_detail,
    get_driver_behavior_export,
)
from app.services.leakage_service import (
    get_leakage_summary,
    get_leakage_drivers,
    get_leakage_export,
)
from app.services.top_driver_behavior_service import (
    get_top_driver_behavior_summary,
    get_top_driver_behavior_benchmarks,
    get_top_driver_behavior_patterns,
    get_top_driver_behavior_playbook_insights,
    get_top_driver_behavior_export,
)
from app.services.business_slice_service import (
    append_unmapped_bucket_rows,
    enrich_business_slice_matrix_meta,
    get_business_slice_filters,
    get_business_slice_monthly,
    get_business_slice_coverage,
    get_business_slice_coverage_summary,
    get_business_slice_unmatched,
    get_business_slice_conflicts,
    get_business_slice_subfleets,
    get_plan_business_slice_stub,
    get_business_slice_weekly,
    get_business_slice_daily,
    get_business_slice_matrix_freshness_meta,
)
from app.services.business_slice_omniview_service import get_business_slice_omniview
from app.services.omniview_matrix_integrity_service import (
    get_matrix_operational_trust_api_payload,
    log_omniview_issue_action,
)
from app.settings import settings
from fastapi.responses import Response
from typing import Optional, Literal
import asyncio
import functools
import csv
import io
import json
import logging
import os
import sys
import time

# Ejecutar función síncrona en thread pool (compatible Python 3.8+; no usar asyncio.to_thread que es 3.9+)
async def _run_sync(func, *args, **kwargs):
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, functools.partial(func, *args, **kwargs))


async def _run_sync_request(request: Request, func, *args, **kwargs):
    """Como _run_sync, pero si el cliente cierra la petición (AbortController, pestaña), llama conn.cancel() en Postgres."""
    from app.db.connection import cancel_pg_queries_for_thread, clear_pg_registrations_for_thread

    state: dict = {"tid": None}

    def _wrapper():
        tid = threading.get_ident()
        state["tid"] = tid
        try:
            return func(*args, **kwargs)
        finally:
            clear_pg_registrations_for_thread(tid)
            state["tid"] = None

    async def _watch_disconnect():
        while True:
            if await request.is_disconnected():
                tid = state.get("tid")
                if tid is not None:
                    cancel_pg_queries_for_thread(tid)
                return
            await asyncio.sleep(0.15)

    watcher = asyncio.create_task(_watch_disconnect())
    loop = asyncio.get_event_loop()
    try:
        return await loop.run_in_executor(None, _wrapper)
    finally:
        watcher.cancel()
        with suppress(asyncio.CancelledError):
            await watcher

logger = logging.getLogger(__name__)
logger.info("ops router: _run_sync uses run_in_executor (Python 3.8 compatible)")
# #region agent log
def _debug_log(location: str, message: str, data: dict, hypothesis_id: str):
    try:
        log_path = os.path.join(os.path.dirname(__file__), "..", "..", "..", "debug-1c8c83.log")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(json.dumps({"sessionId": "1c8c83", "timestamp": int(time.time() * 1000), "location": location, "message": message, "data": data, "hypothesisId": hypothesis_id}) + "\n")
    except Exception:
        pass
# #endregion

router = APIRouter(prefix="/ops", tags=["ops"])

@router.get("/universe")
async def get_universe(
    country: Optional[str] = Query(None, description="Filtrar por país"),
    city: Optional[str] = Query(None, description="Filtrar por ciudad")
):
    """
    Retorna el universo operativo: combinaciones (country, city, line_of_business)
    válidas con actividad real en 2025.
    Soporta filtros opcionales por country y city.
    """
    try:
        universe = get_ops_universe(country=country, city=city)
        return {
            "data": universe,
            "total_combinations": len(universe)
        }
    except Exception as e:
        logger.error(f"Error al obtener universo operativo: {e}")
        raise

@router.get("/territory-quality/kpis")
async def get_territory_quality_kpis(
    granularity: Literal["total", "weekly"] = Query("total", description="Granularidad: total o weekly")
):
    """
    Obtiene KPIs de calidad de mapeo territorial.
    """
    try:
        if granularity == "total":
            kpis = get_territory_kpis_total()
            return {
                "granularity": "total",
                "data": kpis
            }
        else:  # weekly
            kpis = get_territory_kpis_weekly()
            return {
                "granularity": "weekly",
                "data": kpis,
                "total_weeks": len(kpis)
            }
    except Exception as e:
        logger.error(f"Error al obtener KPIs de territorio: {e}")
        raise HTTPException(status_code=500, detail=f"Error al obtener KPIs de territorio: {str(e)}")

@router.get("/territory-quality/unmapped-parks")
async def get_unmapped_parks_endpoint(
    limit: int = Query(50, description="Límite de resultados", ge=1, le=500)
):
    """
    Obtiene parks que aparecen en trips_all pero no tienen mapeo en dim.dim_park.
    Ordenados por cantidad de trips (descendente).
    """
    try:
        parks = get_unmapped_parks(limit=limit)
        return {
            "data": parks,
            "total_parks": len(parks),
            "limit": limit
        }
    except Exception as e:
        logger.error(f"Error al obtener parks unmapped: {e}")
        raise HTTPException(status_code=500, detail=f"Error al obtener parks unmapped: {str(e)}")

def _plan_vs_real_resolve_source(
    source: Optional[str],
    country: Optional[str],
) -> tuple[bool, str, str]:
    """
    Resuelve si usar canonical o legacy.
    - source=canonical → canonical (forzado).
    - source=legacy → legacy (forzado).
    - USE_CANONICAL_PLAN_VS_REAL_DEFAULT y parity MATCH/MINOR → canonical; si MAJOR → fallback a legacy.
    - Resto → legacy.
    Retorna (use_canonical, parity_status, data_completeness).
    """
    src = (source or "").strip().lower()
    if src == "legacy":
        audit = get_latest_parity_audit(scope=country.strip().lower() if country else None)
        parity = (audit or {}).get("diagnosis") or "UNKNOWN"
        completeness = (audit or {}).get("data_completeness") or "FULL"
        return False, parity, completeness
    forced_canonical = src == "canonical"
    if forced_canonical:
        audit = get_latest_parity_audit(scope=country.strip().lower() if country else None)
        parity = (audit or {}).get("diagnosis") or "UNKNOWN"
        completeness = (audit or {}).get("data_completeness") or "FULL"
        return True, parity, completeness

    if not getattr(settings, "USE_CANONICAL_PLAN_VS_REAL_DEFAULT", False):
        return False, "UNKNOWN", "FULL"

    scope = (country or "").strip().lower() or "global"
    audit = get_latest_parity_audit(scope=scope)
    if not audit:
        return False, "UNKNOWN", "FULL"
    diagnosis = (audit.get("diagnosis") or "").upper()
    completeness = audit.get("data_completeness") or "FULL"
    if diagnosis in ("MATCH", "MINOR_DIFF"):
        return True, diagnosis, completeness
    return False, diagnosis, completeness


@router.get("/plan-vs-real/monthly")
async def get_plan_vs_real_monthly_endpoint(
    country: Optional[str] = Query(None, description="Filtrar por país"),
    city: Optional[str] = Query(None, description="Filtrar por ciudad"),
    real_tipo_servicio: Optional[str] = Query(None, description="Filtrar por tipo de servicio (dimensión real)"),
    park_id: Optional[str] = Query(None, description="Filtrar por park_id"),
    month: Optional[str] = Query(None, description="Filtrar por mes (formato: YYYY-MM o YYYY-MM-DD)"),
    source: Optional[str] = Query(None, description="Fuente: 'canonical' para real canónica; omitir = según parity y flag")
):
    """
    Comparación Plan vs Real mensual. source=canonical fuerza canónica; si no, se usa canonical solo cuando
    parity es MATCH o MINOR_DIFF y USE_CANONICAL_PLAN_VS_REAL_DEFAULT=True; si no, legacy.
    """
    use_canonical, parity_status, data_completeness = _plan_vs_real_resolve_source(source, country)
    try:
        data = get_plan_vs_real_monthly(
            country=country,
            city=city,
            real_tipo_servicio=real_tipo_servicio,
            park_id=park_id,
            month=month,
            use_canonical=use_canonical
        )
        if not use_canonical:
            log_plan_vs_real_source_usage(
                "legacy",
                "/ops/plan-vs-real/monthly",
                {"country": country, "city": city, "month": month},
            )
        return {
            "data": data,
            "total_records": len(data),
            "source_status": "canonical" if use_canonical else "legacy",
            "parity_status": parity_status,
            "data_completeness": data_completeness,
        }
    except Exception as e:
        logger.error(f"Error al obtener comparación Plan vs Real: {e}")
        raise HTTPException(status_code=500, detail=f"Error al obtener comparación Plan vs Real: {str(e)}")


@router.get("/control-loop/plan-versions")
async def control_loop_plan_versions_endpoint():
    """Lista `plan_version` disponibles en staging Control Loop."""
    try:
        versions = list_control_loop_plan_versions()
        return {"plan_versions": versions, "total": len(versions)}
    except Exception as e:
        logger.error("control-loop/plan-versions: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/control-loop/plan-vs-real")
async def control_loop_plan_vs_real_endpoint(
    plan_version: str = Query(..., description="Versión de plan cargada (staging Control Loop)"),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    linea_negocio: Optional[str] = Query(None, description="Línea canónica (ej. auto_taxi)"),
    period_from: Optional[str] = Query(None, description="YYYY-MM inicio"),
    period_to: Optional[str] = Query(None, description="YYYY-MM fin"),
):
    """
    Comparación Plan vs Real para la proyección agregada Control Loop.
    No usa Omniview Matrix ni `ops.v_plan_vs_real_realkey_final`.
    """
    try:
        data = get_control_loop_plan_vs_real(
            plan_version=plan_version,
            country=country,
            city=city,
            linea_negocio=linea_negocio,
            period_from=period_from,
            period_to=period_to,
        )
        return sanitize_for_json({"data": data, "total_records": len(data)})
    except Exception as e:
        logger.exception("control-loop/plan-vs-real")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/business-slice/omniview-projection")
async def business_slice_omniview_projection(
    plan_version: str = Query(..., description="Plan version activa (staging Control Loop)"),
    grain: Literal["monthly", "weekly", "daily"] = Query("monthly"),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    business_slice: Optional[str] = Query(None, description="Filtrar por tajada (business_slice_name)"),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    debug_distribution: bool = Query(False, description="Devuelve auditoría del reparto mensual→semanal/diario"),
):
    """
    Omniview Projection Mode — Plan vs Real con curva estacional.
    Devuelve filas compatibles con la matriz Omniview, con métricas de
    cumplimiento (attainment) basadas en expected_to_date no lineal.
    Aditivo: no modifica contratos de /monthly, /weekly, /daily.
    """
    try:
        data = get_omniview_projection(
            plan_version=plan_version,
            grain=grain,
            country=country,
            city=city,
            business_slice=business_slice,
            year=year,
            month=month,
            debug_distribution=debug_distribution,
        )
        return sanitize_for_json(data)
    except Exception as e:
        logger.exception("business-slice/omniview-projection")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/business-slice/real-freshness")
async def business_slice_real_freshness_endpoint():
    """
    Freshness Omniview: upstream (trips base) + facts agregados (day/week/month),
    status global (peor caso), lag combinado, last_refresh_at, next_scheduled_run.
    """
    try:
        return sanitize_for_json(get_omniview_business_slice_real_freshness())
    except Exception as e:
        logger.exception("business-slice/real-freshness")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/business-slice/real-refresh-omniview")
async def business_slice_real_refresh_omniview_endpoint(
    force: bool = Query(False, description="Si True, ignora cooldown entre corridas."),
):
    """
    Ejecuta recarga operacional de day_fact + week_fact (mes actual y anterior).
    Puede tardar minutos; usar desde cron, Task Scheduler o admin.
    """
    try:
        result = await _run_sync(run_business_slice_real_refresh_job, force)
        return sanitize_for_json(result)
    except Exception as e:
        logger.exception("business-slice/real-refresh-omniview")
        raise HTTPException(status_code=500, detail=str(e))


# ───────────────────────────────────────────────────────────────────────
# FASE_KPI_CONSISTENCY: endpoints de auditoría operativa.
# Reutilizan exactamente las funciones que usan los CLI scripts:
#   backend/scripts/validate_kpi_grain_consistency.py
#   backend/scripts/debug_rollup_mismatch.py
# Read-only; no tocan tablas de hechos. Soportan format=json|csv.
# ───────────────────────────────────────────────────────────────────────

def _parse_month_param(month: Optional[str]) -> tuple[int, int]:
    """
    Acepta YYYY-MM o YYYY-MM-DD; si None, usa primer día del mes actual.
    """
    from datetime import date as _date
    if not month:
        d = _date.today().replace(day=1)
        return d.year, d.month
    s = str(month).strip()
    try:
        if len(s) == 7:  # YYYY-MM
            y, m = s.split("-")
            return int(y), int(m)
        d = _date.fromisoformat(s)
        return d.year, d.month
    except Exception as ex:
        raise HTTPException(status_code=400, detail=f"month inválido (esperado YYYY-MM o YYYY-MM-DD): {ex}")


def _csv_response_from_rows(rows: list[dict], columns: list[str], filename: str) -> Response:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=columns)
    w.writeheader()
    for r in rows:
        w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in columns})
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/kpi-consistency-audit")
async def kpi_consistency_audit_endpoint(
    month: Optional[str] = Query(None, description="YYYY-MM o YYYY-MM-DD; default = mes actual."),
    months: int = Query(1, ge=1, le=6, description="Cantidad de meses hacia atrás incluyendo el indicado."),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    format: Literal["json", "csv"] = Query("json"),
):
    """
    Audita la consistencia KPI por grano para el rango [month - months + 1, month]
    aplicando el contrato de `kpi_aggregation_rules.py`.

    Status posibles por celda/KPI:
      - ok                       (additive cuadra)
      - expected_non_comparable  (semi_additive / ratio: comportamiento esperado)
      - warning                  (regla desconocida o dudosa)
      - fail                     (inconsistencia real, requiere acción)
    """
    from scripts.validate_kpi_grain_consistency import (
        run_consistency_audit,
        summarize as _summarize,
        CSV_COLUMNS,
    )
    from datetime import date as _date

    try:
        y, m = _parse_month_param(month)

        def _months_back(n: int) -> list[tuple[int, int]]:
            out: list[tuple[int, int]] = []
            yy, mm = y, m
            for _ in range(n):
                out.append((yy, mm))
                mm -= 1
                if mm == 0:
                    mm = 12
                    yy -= 1
            return list(reversed(out))

        def _do_run() -> dict:
            all_rows: list[dict] = []
            per_month: list[dict] = []
            for yy, mm in _months_back(months):
                rows = run_consistency_audit(yy, mm, country, city)
                summary = _summarize(rows)
                per_month.append({"month": _date(yy, mm, 1).isoformat(), "summary": summary})
                all_rows.extend(rows)
            return {
                "rows": all_rows,
                "per_month": per_month,
                "summary_total": _summarize(all_rows),
            }

        result = await _run_sync(_do_run)

        if format == "csv":
            ts = _date(y, m, 1).isoformat()
            return _csv_response_from_rows(
                result["rows"],
                CSV_COLUMNS,
                f"kpi_consistency_audit_{ts}_back{months}.csv",
            )

        return sanitize_for_json({
            "generated_at": _time.time(),
            "params": {
                "month": _date(y, m, 1).isoformat(),
                "months": months,
                "country": country,
                "city": city,
            },
            "summary": result["summary_total"],
            "per_month": result["per_month"],
            "rows": result["rows"],
        })
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("kpi-consistency-audit")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/rollup-mismatch-audit")
async def rollup_mismatch_audit_endpoint(
    month: str = Query(..., description="YYYY-MM o YYYY-MM-DD."),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    business_slice: Optional[str] = Query(None),
    include_resolved: bool = Query(False, description="Si True, compara también vs v_real_trips_business_slice_resolved (más lento)."),
    format: Literal["json", "csv"] = Query("json"),
):
    """
    Diagnóstico fuerte de ROLLUP_MISMATCH para un mes:
    compara month_fact vs SUM(day_fact) y, opcionalmente, vs v_real_trips_business_slice_resolved.
    Cada celda lleva un `suspected_cause` (stale_month_fact, stale_day_fact,
    duplication_or_mapping, filter_mismatch_vs_resolved, mapping_mismatch_*, negligible, etc.).
    """
    from scripts.debug_rollup_mismatch import (
        run_audit,
        summarize as _summarize,
        CSV_COLUMNS,
    )
    from datetime import date as _date

    try:
        y, m = _parse_month_param(month)

        def _do_run() -> dict:
            rows = run_audit(y, m, country, city, business_slice, include_resolved)
            return {"rows": rows, "summary": _summarize(rows)}

        result = await _run_sync(_do_run)

        if format == "csv":
            return _csv_response_from_rows(
                result["rows"],
                CSV_COLUMNS,
                f"rollup_mismatch_audit_{_date(y, m, 1).isoformat()}.csv",
            )

        return sanitize_for_json({
            "generated_at": _time.time(),
            "params": {
                "month": _date(y, m, 1).isoformat(),
                "country": country,
                "city": city,
                "business_slice": business_slice,
                "include_resolved": include_resolved,
            },
            "summary": result["summary"],
            "rows": result["rows"],
        })
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("rollup-mismatch-audit")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/decision-readiness")
async def decision_readiness_endpoint(
    format: Literal["json", "csv"] = Query("json"),
):
    """
    FASE_VALIDATION_FIX: reporte de decision readiness por KPI.
    Devuelve el estado de cada KPI respecto a su usabilidad para decisiones
    ejecutivas: decision_ready | scope_only | formula_only | restricted.

    Sin parámetros de filtro — es estático (deriva del contrato KPI en memoria).
    """
    from scripts.report_decision_readiness import (
        build_decision_readiness_rows,
        CSV_COLUMNS as _DR_CSV_COLUMNS,
    )

    try:
        rows = build_decision_readiness_rows()

        summary: dict[str, list[str]] = {}
        for r in rows:
            s = r["decision_status"]
            summary.setdefault(s, []).append(r["kpi"])

        if format == "csv":
            return _csv_response_from_rows(rows, _DR_CSV_COLUMNS, "decision_readiness.csv")

        return sanitize_for_json({
            "generated_at": _time.time(),
            "summary": summary,
            "rows": rows,
        })
    except Exception as e:
        logger.exception("decision-readiness")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/plan-vs-real/alerts")
async def get_plan_vs_real_alerts_endpoint(
    country: Optional[str] = Query(None, description="Filtrar por país"),
    month: Optional[str] = Query(None, description="Filtrar por mes (formato: YYYY-MM o YYYY-MM-DD)"),
    alert_level: Optional[str] = Query(None, description="Filtrar por nivel de alerta (CRITICO, MEDIO, OK)"),
    source: Optional[str] = Query(None, description="Fuente: 'canonical' para real canónica; omitir = según parity")
):
    """
    Alertas Plan vs Real (solo matched). Misma lógica de source/parity que /plan-vs-real/monthly.
    """
    use_canonical, parity_status, data_completeness = _plan_vs_real_resolve_source(source, country)
    try:
        data = get_alerts_monthly(
            country=country,
            month=month,
            alert_level=alert_level,
            use_canonical=use_canonical
        )
        if not use_canonical:
            log_plan_vs_real_source_usage(
                "legacy",
                "/ops/plan-vs-real/alerts",
                {"country": country, "month": month, "alert_level": alert_level},
            )
        return {
            "data": data,
            "total_alerts": len(data),
            "source_status": "canonical" if use_canonical else "legacy",
            "parity_status": parity_status,
            "data_completeness": data_completeness,
            "by_level": {
                "CRITICO": len([a for a in data if a.get("alert_level") == "CRITICO"]),
                "MEDIO": len([a for a in data if a.get("alert_level") == "MEDIO"]),
                "OK": len([a for a in data if a.get("alert_level") == "OK"])
            }
        }
    except Exception as e:
        logger.error(f"Error al obtener alertas Plan vs Real: {e}")
        raise HTTPException(status_code=500, detail=f"Error al obtener alertas Plan vs Real: {str(e)}")

def _real_lob_meta():
    meta = get_real_lob_meta()
    return {
        "last_available_month": meta.get("max_month"),
        "last_available_week": meta.get("max_week"),
    }


@router.get("/real-lob/monthly")
async def get_real_lob_monthly_endpoint(
    country: Optional[str] = Query(None, description="Filtrar por país"),
    city: Optional[str] = Query(None, description="Filtrar por ciudad"),
    lob_name: Optional[str] = Query(None, description="Filtrar por LOB"),
    month: Optional[str] = Query(None, description="Filtrar por mes YYYY-MM; si vacío, último mes disponible"),
    year_real: Optional[int] = Query(None, description="Filtrar por año (todos los meses del año)")
):
    """
    REAL LOB Observability: viajes REAL agregados por LOB (mensual).
    Sin month ni year_real: devuelve el último mes disponible.
    """
    try:
        data = get_real_lob_monthly_svc(
            country=country, city=city, lob_name=lob_name, month=month, year_real=year_real
        )
        meta = _real_lob_meta()
        resp = {"data": data, "total_records": len(data), **meta}
        if not data:
            resp["reason"] = "no_data_for_filters"
        return resp
    except Exception as e:
        logger.error(f"Error Real LOB monthly: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-lob/weekly")
async def get_real_lob_weekly_endpoint(
    country: Optional[str] = Query(None, description="Filtrar por país"),
    city: Optional[str] = Query(None, description="Filtrar por ciudad"),
    lob_name: Optional[str] = Query(None, description="Filtrar por LOB"),
    week_start: Optional[str] = Query(None, description="Lunes de la semana YYYY-MM-DD; si vacío, última semana disponible"),
    year_real: Optional[int] = Query(None, description="Filtrar por año (todas las semanas del año)")
):
    """
    REAL LOB Observability: viajes REAL agregados por LOB (semanal).
    Sin week_start ni year_real: devuelve la última semana disponible.
    """
    try:
        data = get_real_lob_weekly_svc(
            country=country, city=city, lob_name=lob_name, week_start=week_start, year_real=year_real
        )
        meta = _real_lob_meta()
        resp = {"data": data, "total_records": len(data), **meta}
        if not data:
            resp["reason"] = "no_data_for_filters"
        return resp
    except Exception as e:
        logger.error(f"Error Real LOB weekly: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-lob/debug")
async def get_real_lob_debug_endpoint():
    """
    Devuelve max_month, max_week, count_month, count_week. Solo disponible en entorno dev.
    """
    env = (getattr(settings, "ENVIRONMENT", "") or "").lower()
    if env not in ("dev", "development"):
        raise HTTPException(status_code=404, detail="Solo disponible en desarrollo")
    try:
        return get_real_lob_meta()
    except Exception as e:
        logger.error(f"Error Real LOB debug: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def _real_lob_meta_v2():
    meta = get_real_lob_meta_v2()
    return {"last_available_month": meta.get("max_month"), "last_available_week": meta.get("max_week")}


@router.get("/real-lob/monthly-v2")
async def get_real_lob_monthly_v2_endpoint(
    country: Optional[str] = Query(None, description="Filtrar por país"),
    city: Optional[str] = Query(None, description="Filtrar por ciudad"),
    park_id: Optional[str] = Query(None, description="Filtrar por park_id"),
    lob_group: Optional[str] = Query(None, description="LOB_GROUP: auto taxi, delivery, tuk tuk, taxi moto, UNCLASSIFIED"),
    real_tipo_servicio: Optional[str] = Query(None, description="Tipo de servicio normalizado"),
    segment_tag: Optional[str] = Query(None, description="Segmento: B2B o B2C"),
    month: Optional[str] = Query(None, description="Mes YYYY-MM; si vacío, último mes"),
    year_real: Optional[int] = Query(None, description="Año (rango de meses)")
):
    """Real LOB v2: mensual con filtros country, city, park_id, lob_group, real_tipo_servicio, segment_tag."""
    try:
        data = get_real_lob_monthly_v2(
            country=country, city=city, park_id=park_id,
            lob_group=lob_group, real_tipo_servicio=real_tipo_servicio, segment_tag=segment_tag,
            month=month, year_real=year_real
        )
        meta = _real_lob_meta_v2()
        resp = {"data": data, "total_records": len(data), **meta}
        if not data:
            resp["reason"] = "no_data_for_filters"
        return resp
    except Exception as e:
        logger.error(f"Error Real LOB monthly v2: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-lob/weekly-v2")
async def get_real_lob_weekly_v2_endpoint(
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    lob_group: Optional[str] = Query(None),
    real_tipo_servicio: Optional[str] = Query(None),
    segment_tag: Optional[str] = Query(None),
    week_start: Optional[str] = Query(None, description="Lunes semana YYYY-MM-DD"),
    year_real: Optional[int] = Query(None)
):
    """Real LOB v2: semanal con mismos filtros."""
    try:
        data = get_real_lob_weekly_v2(
            country=country, city=city, park_id=park_id,
            lob_group=lob_group, real_tipo_servicio=real_tipo_servicio, segment_tag=segment_tag,
            week_start=week_start, year_real=year_real
        )
        meta = _real_lob_meta_v2()
        resp = {"data": data, "total_records": len(data), **meta}
        if not data:
            resp["reason"] = "no_data_for_filters"
        return resp
    except Exception as e:
        logger.error(f"Error Real LOB weekly v2: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/parks")
async def get_ops_parks(
    country: Optional[str] = Query(None, description="Filtrar por país (mismo criterio que Real LOB)"),
    city: Optional[str] = Query(None, description="Filtrar por ciudad"),
):
    """
    Lista de parks para dropdowns (Driver Lifecycle, etc.).
    Misma fuente y orden que Real LOB: [{ park_id, park_name }].
    La opción "Todos" se arma en frontend.
    """
    try:
        filters = get_real_lob_filters(country=country, city=city)
        parks = filters.get("parks") or []
        return {
            "parks": [
                {"park_id": p.get("park_id"), "park_name": p.get("park_name") or str(p.get("park_id") or "")}
                for p in parks
            ]
        }
    except Exception as e:
        logger.error("GET /ops/parks: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# ─── Control Tower Supply (REAL) ─────────────────────────────────────────────
@router.get("/supply/geo")
async def get_supply_geo_endpoint(
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
):
    """Geo para filtros: countries, cities (por country), parks (por country/city). Fuente: dim.v_geo_park."""
    try:
        return get_supply_geo(country=country, city=city)
    except Exception as e:
        logger.error("GET /ops/supply/geo: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/supply/parks")
async def get_supply_parks_endpoint(
    country: Optional[str] = Query(None, description="Filtrar por país"),
    city: Optional[str] = Query(None, description="Filtrar por ciudad"),
):
    """Parks para Supply (geo). Fuente: dim.v_geo_park. Orden: country, city, park_name."""
    try:
        data = get_supply_parks(country=country, city=city)
        return {"data": data}
    except Exception as e:
        logger.error("GET /ops/supply/parks: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


def _to_csv(rows: list, columns: list) -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(columns)
    for r in rows:
        w.writerow([r.get(c) for c in columns])
    return buf.getvalue()


@router.get("/supply/series")
async def get_supply_series_endpoint(
    park_id: str = Query(..., description="Park (obligatorio)"),
    from_: str = Query(..., alias="from", description="Fecha inicio YYYY-MM-DD"),
    to: str = Query(..., description="Fecha fin YYYY-MM-DD"),
    grain: Literal["weekly", "monthly"] = Query("weekly"),
    format: Optional[str] = Query(None, description="csv para descarga"),
):
    """Serie por periodo (DESC). park_id obligatorio."""
    try:
        data = get_supply_series(park_id=park_id, from_date=from_, to_date=to, grain=grain)
        if (format or "").lower() == "csv":
            # Regla presentación: no IDs en export; solo columnas legibles (park_name, city, country)
            cols = ["period_start", "park_name", "city", "country", "activations", "active_drivers", "churned", "reactivated", "churn_rate", "reactivation_rate", "net_growth"]
            body = _to_csv(data, [c for c in cols if data and data[0].get(c) is not None] or cols)
            return Response(content=body, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=supply_series.csv"})
        return {"data": data}
    except Exception as e:
        logger.error("GET /ops/supply/series: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/supply/segments/series")
async def get_supply_segments_series_endpoint(
    park_id: str = Query(..., description="Park (obligatorio)"),
    from_: str = Query(..., alias="from", description="YYYY-MM-DD"),
    to: str = Query(..., description="YYYY-MM-DD"),
    format: Optional[str] = Query(None),
):
    """Serie de segmentos por semana (FT/PT/CASUAL/OCC/DORMANT). Fuente: ops.mv_supply_segments_weekly. Orden: week_start DESC."""
    try:
        data = get_supply_segments_series(park_id=park_id, from_date=from_, to_date=to)
        if (format or "").lower() == "csv":
            cols = ["week_start", "segment_week", "drivers_count", "trips_sum", "share_of_active", "park_name", "city", "country"]
            body = _to_csv(data, cols)
            return Response(
                content=body,
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=supply_segments_series.csv"},
            )
        return {"data": data}
    except Exception as e:
        logger.error("GET /ops/supply/segments/series: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/supply/segments/config")
async def get_supply_segment_config_endpoint():
    """Configuración de segmentos (ops.driver_segment_config): segment, min_trips, max_trips, priority. Sustituye umbrales hardcodeados."""
    try:
        return {"data": get_supply_segment_config()}
    except Exception as e:
        logger.error("GET /ops/supply/segments/config: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/supply/summary")
async def get_supply_summary_endpoint(
    park_id: str = Query(..., description="Park (obligatorio)"),
    from_: str = Query(..., alias="from"),
    to: str = Query(..., description="Fecha fin YYYY-MM-DD"),
    grain: Literal["weekly", "monthly"] = Query("weekly"),
):
    """Summary cards del rango visible (sumas y tasas ponderadas)."""
    try:
        return get_supply_summary(park_id=park_id, from_date=from_, to_date=to, grain=grain)
    except Exception as e:
        logger.error("GET /ops/supply/summary: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/supply/global/series")
async def get_supply_global_series_endpoint(
    from_: str = Query(..., alias="from"),
    to: str = Query(..., description="Fecha fin YYYY-MM-DD"),
    grain: Literal["weekly", "monthly"] = Query("weekly"),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    format: Optional[str] = Query(None),
):
    """Serie global (agregada por periodo; opcional country/city)."""
    try:
        data = get_supply_global_series(from_date=from_, to_date=to, grain=grain, country=country, city=city)
        if (format or "").lower() == "csv":
            cols = ["period_start", "activations", "active_drivers", "churned", "reactivated", "net_growth"]
            if data and any("country" in r for r in data):
                cols = ["period_start", "country", "city"] + [c for c in cols if c != "period_start"]
            body = _to_csv(data, cols)
            return Response(content=body, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=supply_global.csv"})
        return {"data": data}
    except Exception as e:
        logger.error("GET /ops/supply/global/series: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/supply/alerts")
async def get_supply_alerts_endpoint(
    park_id: Optional[str] = Query(None),
    from_: Optional[str] = Query(None, alias="from", description="Semana desde YYYY-MM-DD"),
    to: Optional[str] = Query(None, description="Semana hasta YYYY-MM-DD"),
    week_start_from: Optional[str] = Query(None),
    week_start_to: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    alert_type: Optional[str] = Query(None, description="segment_drop | segment_spike"),
    severity: Optional[str] = Query(None, description="P0 | P1 | P2 | P3"),
    limit: int = Query(200, ge=1, le=500),
    format: Optional[str] = Query(None),
):
    """Alertas Supply PRO por semana, park, segmento. Fuente: ops.mv_supply_alerts_weekly."""
    try:
        from_val = from_ or week_start_from
        to_val = to or week_start_to
        data = get_supply_alerts(
            week_start_from=from_val,
            week_start_to=to_val,
            park_id=park_id,
            country=country,
            city=city,
            alert_type=alert_type,
            severity=severity,
            limit=limit,
        )
        if (format or "").lower() == "csv":
            cols = ["week_start", "severity", "alert_type", "segment_week", "current_value", "baseline_avg", "delta_pct", "message_short", "recommended_action", "park_name", "city", "country"]
            body = _to_csv(data, cols)
            return Response(
                content=body,
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=supply_alerts.csv"},
            )
        return {"data": data, "total": len(data)}
    except Exception as e:
        logger.error("GET /ops/supply/alerts: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/supply/alerts/drilldown")
async def get_supply_alert_drilldown_endpoint(
    park_id: str = Query(..., description="Park ID"),
    week_start: str = Query(..., description="Semana (YYYY-MM-DD)"),
    segment_week: Optional[str] = Query(None, description="FT | PT | CASUAL | OCCASIONAL"),
    alert_type: Optional[str] = Query(None, description="segment_drop | segment_spike"),
    format: Optional[str] = Query(None, description="csv para export"),
):
    """Conductores afectados (downshift/drop) para una alerta. Orden: baseline_trips_4w_avg desc."""
    try:
        data = get_supply_alert_drilldown(
            week_start=week_start,
            park_id=park_id,
            segment_week=segment_week,
            alert_type=alert_type,
        )
        if (format or "").lower() == "csv":
            cols = ["driver_key", "prev_segment_week", "segment_week_current", "trips_completed_week", "baseline_trips_4w_avg", "segment_change_type", "week_start", "park_id"]
            body = _to_csv(data, cols)
            return Response(
                content=body,
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=supply_alert_drilldown_{week_start}_{park_id or 'all'}.csv"},
            )
        return {"data": data, "total": len(data)}
    except Exception as e:
        logger.error("GET /ops/supply/alerts/drilldown: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/supply/overview-enhanced")
async def get_supply_overview_enhanced_endpoint(
    park_id: str = Query(..., description="Park (obligatorio)"),
    from_: str = Query(..., alias="from", description="YYYY-MM-DD"),
    to: str = Query(..., description="YYYY-MM-DD"),
    grain: Literal["weekly", "monthly"] = Query("weekly"),
):
    """Overview enriquecido: trips, avg_trips_per_driver, FT/PT/weak_supply share; WoW cuando grain=weekly."""
    try:
        data = get_supply_overview_enhanced(park_id=park_id, from_date=from_, to_date=to, grain=grain)
        return data
    except Exception as e:
        logger.error("GET /ops/supply/overview-enhanced: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/supply/composition")
async def get_supply_composition_endpoint(
    park_id: str = Query(..., description="Park (obligatorio)"),
    from_: str = Query(..., alias="from", description="YYYY-MM-DD"),
    to: str = Query(..., description="YYYY-MM-DD"),
    format: Optional[str] = Query(None),
):
    """Composición semanal por segmento con WoW. Fuente: ops.mv_supply_segments_weekly."""
    try:
        data = get_supply_composition(park_id=park_id, from_date=from_, to_date=to)
        if (format or "").lower() == "csv":
            cols = ["week_start", "segment_week", "drivers_count", "delta_drivers", "trips_sum", "share_of_active", "delta_share", "supply_contribution", "avg_trips_per_driver", "drivers_wow_pct", "trips_wow_pct", "share_wow_pp"]
            body = _to_csv(data, [c for c in cols if data and data[0].get(c) is not None] or cols)
            return Response(content=body, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=supply_composition.csv"})
        return {"data": data, "total": len(data)}
    except Exception as e:
        logger.error("GET /ops/supply/composition: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/supply/migration")
async def get_supply_migration_endpoint(
    park_id: str = Query(..., description="Park (obligatorio)"),
    from_: str = Query(..., alias="from", description="YYYY-MM-DD"),
    to: str = Query(..., description="YYYY-MM-DD"),
    format: Optional[str] = Query(None),
):
    """Migración entre segmentos por semana: from_segment, to_segment, drivers_migrated, migration_type. Incluye summary (upgrades, downgrades, drops, revivals)."""
    try:
        result = get_supply_migration(park_id=park_id, from_date=from_, to_date=to)
        rows = result.get("data", [])
        summary = result.get("summary", {"upgrades": 0, "downgrades": 0, "drops": 0, "revivals": 0})
        if (format or "").lower() == "csv":
            cols = ["week_start", "park_id", "from_segment", "to_segment", "migration_type", "drivers_migrated"]
            body = _to_csv(rows, cols)
            return Response(content=body, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=supply_migration.csv"})
        return {"data": rows, "total": len(rows), "summary": summary}
    except Exception as e:
        logger.error("GET /ops/supply/migration: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/supply/migration/drilldown")
async def get_supply_migration_drilldown_endpoint(
    park_id: str = Query(..., description="Park ID"),
    week_start: str = Query(..., description="Semana YYYY-MM-DD"),
    from_segment: Optional[str] = Query(None),
    to_segment: Optional[str] = Query(None),
    format: Optional[str] = Query(None),
):
    """Drivers que migraron en una semana (opcional: from_segment, to_segment)."""
    try:
        data = get_supply_migration_drilldown(park_id=park_id, week_start=week_start, from_segment=from_segment, to_segment=to_segment)
        if (format or "").lower() == "csv":
            cols = ["driver_key", "week_start", "park_id", "from_segment", "to_segment", "migration_type", "trips_completed_week", "baseline_trips_4w_avg"]
            body = _to_csv(data, cols)
            return Response(content=body, media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=supply_migration_drilldown_{week_start}_{park_id}.csv"})
        return {"data": data, "total": len(data)}
    except Exception as e:
        logger.error("GET /ops/supply/migration/drilldown: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/supply/migration/weekly-summary")
async def get_supply_migration_weekly_summary_endpoint(
    park_id: str = Query(..., description="Park (obligatorio)"),
    from_: str = Query(..., alias="from", description="YYYY-MM-DD"),
    to: str = Query(..., description="YYYY-MM-DD"),
):
    """Resumen semanal por segmento (time-first): week_label, segment, drivers, wow_delta, wow_percent, upgrades, downgrades. Requiere vista ops.v_driver_segments_weekly_summary (migración 079)."""
    try:
        data = get_supply_migration_weekly_summary(park_id=park_id, from_date=from_, to_date=to)
        return {"data": data, "total": len(data)}
    except Exception as e:
        logger.error("GET /ops/supply/migration/weekly-summary: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/supply/migration/critical")
async def get_supply_migration_critical_endpoint(
    park_id: str = Query(..., description="Park (obligatorio)"),
    from_: str = Query(..., alias="from", description="YYYY-MM-DD"),
    to: str = Query(..., description="YYYY-MM-DD"),
):
    """Movimientos críticos (drivers > 100 o rate > 15%). Requiere vista ops.v_driver_segment_critical_movements (migración 079)."""
    try:
        data = get_supply_migration_critical(park_id=park_id, from_date=from_, to_date=to)
        return {"data": data, "total": len(data)}
    except Exception as e:
        logger.error("GET /ops/supply/migration/critical: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/supply/definitions")
async def get_supply_definitions_endpoint():
    """Definiciones oficiales de métricas (active_supply, churned, reactivated, growth_rate, segments, migration)."""
    try:
        return get_definitions()
    except Exception as e:
        logger.error("GET /ops/supply/definitions: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/supply/freshness")
async def get_supply_freshness_endpoint():
    """Última semana disponible, última corrida de refresh y estado del pipeline (fresh/stale/unknown)."""
    try:
        return get_supply_freshness()
    except Exception as e:
        logger.error("GET /ops/supply/freshness: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# --- Behavioral Alerts (driver-level baseline deviation) ---
@router.get("/behavior-alerts/summary")
async def get_behavior_alerts_summary_endpoint(
    week_start: Optional[str] = Query(None, description="Semana YYYY-MM-DD"),
    from_: Optional[str] = Query(None, alias="from", description="Desde YYYY-MM-DD"),
    to: Optional[str] = Query(None, description="Hasta YYYY-MM-DD"),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    segment_current: Optional[str] = Query(None),
    movement_type: Optional[str] = Query(None),
    alert_type: Optional[str] = Query(None),
    severity: Optional[str] = Query(None),
    risk_band: Optional[str] = Query(None),
):
    """KPIs: drivers_monitored, critical_drops, moderate_drops, strong_recoveries, silent_erosion, high_volatility, high_risk_drivers, medium_risk_drivers."""
    try:
        return get_behavior_alerts_summary(
            week_start=week_start, from_date=from_, to_date=to,
            country=country, city=city, park_id=park_id,
            segment_current=segment_current, movement_type=movement_type,
            alert_type=alert_type, severity=severity, risk_band=risk_band,
        )
    except Exception as e:
        logger.error("GET /ops/behavior-alerts/summary: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/behavior-alerts/insight")
async def get_behavior_alerts_insight_endpoint(
    week_start: Optional[str] = Query(None),
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    segment_current: Optional[str] = Query(None),
    movement_type: Optional[str] = Query(None),
    alert_type: Optional[str] = Query(None),
    severity: Optional[str] = Query(None),
    risk_band: Optional[str] = Query(None),
):
    """Resumen de texto para el panel de insight (alertas y riesgo)."""
    try:
        return get_behavior_alerts_insight(
            week_start=week_start, from_date=from_, to_date=to,
            country=country, city=city, park_id=park_id,
            segment_current=segment_current, movement_type=movement_type,
            alert_type=alert_type, severity=severity, risk_band=risk_band,
        )
    except Exception as e:
        logger.error("GET /ops/behavior-alerts/insight: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/behavior-alerts/drivers")
async def get_behavior_alerts_drivers_endpoint(
    week_start: Optional[str] = Query(None),
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    segment_current: Optional[str] = Query(None),
    movement_type: Optional[str] = Query(None),
    alert_type: Optional[str] = Query(None),
    severity: Optional[str] = Query(None),
    risk_band: Optional[str] = Query(None),
    limit: int = Query(500, ge=1, le=2000),
    offset: int = Query(0, ge=0),
    order_by: str = Query("risk_score", description="risk_score | severity | delta_pct | week_start"),
    order_dir: str = Query("desc", description="asc | desc"),
):
    """Lista de alertas por conductor (tabla). Orden por defecto: risk_score desc, delta_pct asc."""
    try:
        return get_behavior_alerts_drivers(
            week_start=week_start, from_date=from_, to_date=to,
            country=country, city=city, park_id=park_id,
            segment_current=segment_current, movement_type=movement_type,
            alert_type=alert_type, severity=severity, risk_band=risk_band,
            limit=limit, offset=offset, order_by=order_by, order_dir=order_dir,
        )
    except Exception as e:
        logger.error("GET /ops/behavior-alerts/drivers: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/behavior-alerts/driver-detail")
async def get_behavior_alerts_driver_detail_endpoint(
    driver_key: str = Query(..., description="ID del conductor"),
    week_start: Optional[str] = Query(None),
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = Query(None),
    weeks: int = Query(8, ge=1, le=24),
):
    """Timeline del conductor: viajes últimas N semanas, segmento, baseline, desviación, alerta."""
    try:
        return get_behavior_alerts_driver_detail(
            driver_key=driver_key, week_start=week_start, from_date=from_, to_date=to, weeks=weeks,
        )
    except Exception as e:
        logger.error("GET /ops/behavior-alerts/driver-detail: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/behavior-alerts/export")
async def get_behavior_alerts_export_endpoint(
    week_start: Optional[str] = Query(None),
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    segment_current: Optional[str] = Query(None),
    movement_type: Optional[str] = Query(None),
    alert_type: Optional[str] = Query(None),
    severity: Optional[str] = Query(None),
    risk_band: Optional[str] = Query(None),
    format: Optional[str] = Query("csv", description="csv | excel"),
    max_rows: int = Query(10000, ge=1, le=50000),
):
    """Exporta alertas con filtros activos. format: csv (default) o excel."""
    try:
        rows = get_behavior_alerts_export(
            week_start=week_start, from_date=from_, to_date=to,
            country=country, city=city, park_id=park_id,
            segment_current=segment_current, movement_type=movement_type,
            alert_type=alert_type, severity=severity, risk_band=risk_band,
            max_rows=max_rows,
        )
        cols = ["driver_key", "driver_name", "country", "city", "park_name", "week_label", "segment_current", "movement_type", "trips_current_week", "avg_trips_baseline", "delta_abs", "delta_pct", "weeks_declining_consecutively", "weeks_rising_consecutively", "alert_type", "alert_severity", "risk_score", "risk_band", "last_trip_date"]
        if (format or "csv").lower() == "excel":
            try:
                import openpyxl
                from io import BytesIO
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Behavioral Alerts"
                ws.append(cols)
                for r in rows:
                    ws.append([r.get(c) for c in cols])
                buf = BytesIO()
                wb.save(buf)
                buf.seek(0)
                return Response(
                    content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=behavior_alerts.xlsx"},
                )
            except ImportError:
                body = _to_csv(rows, cols)
                return Response(content=body, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=behavior_alerts.csv"})
        body = _to_csv(rows, cols)
        return Response(content=body, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=behavior_alerts.csv"})
    except Exception as e:
        logger.error("GET /ops/behavior-alerts/export: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# --- Fleet Leakage Monitor MVP ---
@router.get("/leakage/summary")
async def get_leakage_summary_endpoint(
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    leakage_status: Optional[str] = Query(None),
    recovery_priority: Optional[str] = Query(None),
    top_performers_only: Optional[bool] = Query(None),
):
    """KPIs: drivers_under_watch, progressive_leakage, lost_drivers, top_performers_at_risk, cohort_retention_45d."""
    try:
        return get_leakage_summary(
            country=country, city=city, park_id=park_id,
            leakage_status=leakage_status, recovery_priority=recovery_priority,
            top_performers_only=top_performers_only,
        )
    except Exception as e:
        logger.error("GET /ops/leakage/summary: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/leakage/drivers")
async def get_leakage_drivers_endpoint(
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    leakage_status: Optional[str] = Query(None),
    recovery_priority: Optional[str] = Query(None),
    top_performers_only: Optional[bool] = Query(None),
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    order_by: str = Query("leakage_score"),
    order_dir: str = Query("desc"),
):
    """Lista de conductores con clasificación de leakage (stable_retained, watchlist, progressive_leakage, lost_driver)."""
    try:
        return get_leakage_drivers(
            country=country, city=city, park_id=park_id,
            leakage_status=leakage_status, recovery_priority=recovery_priority,
            top_performers_only=top_performers_only,
            limit=limit, offset=offset, order_by=order_by, order_dir=order_dir,
        )
    except Exception as e:
        logger.error("GET /ops/leakage/drivers: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/leakage/export")
async def get_leakage_export_endpoint(
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    leakage_status: Optional[str] = Query(None),
    recovery_priority: Optional[str] = Query(None),
    top_performers_only: Optional[bool] = Query(None),
    format: Optional[str] = Query("csv"),
    max_rows: int = Query(10000, ge=1, le=50000),
):
    """Export Recovery Queue CSV (o Excel si format=excel)."""
    try:
        rows = get_leakage_export(
            country=country, city=city, park_id=park_id,
            leakage_status=leakage_status, recovery_priority=recovery_priority,
            top_performers_only=top_performers_only,
            max_rows=max_rows,
        )
        cols = ["driver_key", "driver_name", "country", "city", "park_name", "segment_week", "trips_current_week",
                "baseline_trips_4w_avg", "delta_pct", "last_trip_date", "days_since_last_trip",
                "leakage_status", "leakage_score", "recovery_priority", "top_performer_at_risk"]
        if (format or "csv").lower() == "excel":
            try:
                import openpyxl
                from io import BytesIO
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Fleet Leakage"
                ws.append(cols)
                for r in rows:
                    ws.append([r.get(c) for c in cols])
                buf = BytesIO()
                wb.save(buf)
                buf.seek(0)
                return Response(
                    content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=fleet_leakage.xlsx"},
                )
            except ImportError:
                body = _to_csv(rows, cols)
                return Response(content=body, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=fleet_leakage.csv"})
        body = _to_csv(rows, cols)
        return Response(content=body, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=fleet_leakage.csv"})
    except Exception as e:
        logger.error("GET /ops/leakage/export: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# --- Action Engine (cohorts + recommended actions) ---
@router.get("/action-engine/summary")
async def get_action_engine_summary_endpoint(
    week_start: Optional[str] = Query(None),
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    segment_current: Optional[str] = Query(None),
    cohort_type: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
):
    """KPIs: actionable_drivers, cohorts_detected, high_priority_cohorts, recoverable_drivers, high_value_at_risk, near_upgrade_opportunities."""
    try:
        return await _run_sync(
            get_action_engine_summary,
            week_start=week_start, from_date=from_, to_date=to,
            country=country, city=city, park_id=park_id,
            segment_current=segment_current, cohort_type=cohort_type, priority=priority,
        )
    except Exception as e:
        logger.error("GET /ops/action-engine/summary: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/action-engine/cohorts")
async def get_action_engine_cohorts_endpoint(
    week_start: Optional[str] = Query(None),
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    segment_current: Optional[str] = Query(None),
    cohort_type: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
):
    """Lista de cohortes con filtros."""
    try:
        return await _run_sync(
            get_action_engine_cohorts,
            week_start=week_start, from_date=from_, to_date=to,
            country=country, city=city, park_id=park_id,
            segment_current=segment_current, cohort_type=cohort_type, priority=priority,
            limit=limit, offset=offset,
        )
    except Exception as e:
        logger.error("GET /ops/action-engine/cohorts: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/action-engine/cohort-detail")
async def get_action_engine_cohort_detail_endpoint(
    cohort_type: str = Query(..., description="Tipo de cohorte"),
    week_start: str = Query(..., description="Semana YYYY-MM-DD"),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    limit: int = Query(500, ge=1, le=2000),
    offset: int = Query(0, ge=0),
):
    """Drivers en una cohorte para drilldown y export."""
    try:
        return await _run_sync(
            get_action_engine_cohort_detail,
            cohort_type=cohort_type, week_start=week_start,
            country=country, city=city, park_id=park_id,
            limit=limit, offset=offset,
        )
    except Exception as e:
        logger.error("GET /ops/action-engine/cohort-detail: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/action-engine/recommendations")
async def get_action_engine_recommendations_endpoint(
    week_start: Optional[str] = Query(None),
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    segment_current: Optional[str] = Query(None),
    top_n: int = Query(5, ge=1, le=20),
):
    """Top N acciones recomendadas para el panel."""
    try:
        return await _run_sync(
            get_action_engine_recommendations,
            week_start=week_start, from_date=from_, to_date=to,
            country=country, city=city, park_id=park_id,
            segment_current=segment_current, top_n=top_n,
        )
    except Exception as e:
        logger.error("GET /ops/action-engine/recommendations: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/action-engine/export")
async def get_action_engine_export_endpoint(
    week_start: Optional[str] = Query(None),
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    segment_current: Optional[str] = Query(None),
    cohort_type: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    format: Optional[str] = Query("csv", description="csv | excel"),
    max_rows: int = Query(10000, ge=1, le=50000),
):
    """Exporta conductores accionables con filtros activos."""
    try:
        rows = await _run_sync(
            get_action_engine_export,
            week_start=week_start, from_date=from_, to_date=to,
            country=country, city=city, park_id=park_id,
            segment_current=segment_current, cohort_type=cohort_type, priority=priority,
            max_rows=max_rows,
        )
        cols = ["driver_key", "driver_name", "week_start", "week_label", "country", "city", "park_id", "park_name", "segment_current", "segment_previous", "movement_type", "trips_current_week", "avg_trips_baseline", "delta_abs", "delta_pct", "alert_type", "severity", "risk_score", "risk_band", "cohort_type"]
        if (format or "csv").lower() == "excel":
            try:
                import openpyxl
                from io import BytesIO
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Action Engine"
                ws.append(cols)
                for r in rows:
                    ws.append([r.get(c) for c in cols])
                buf = BytesIO()
                wb.save(buf)
                return Response(
                    content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=action_engine.xlsx"},
                )
            except ImportError:
                pass
        body = _to_csv(rows, cols)
        return Response(content=body, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=action_engine.csv"})
    except Exception as e:
        logger.error("GET /ops/action-engine/export: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# --- Driver Behavior (deviation engine: time windows, days_since_last_trip) ---
@router.get("/driver-behavior/summary")
async def get_driver_behavior_summary_endpoint(
    recent_weeks: int = Query(4, ge=1, le=32),
    baseline_weeks: int = Query(16, ge=1, le=32),
    as_of_week: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    segment_current: Optional[str] = Query(None),
    alert_type: Optional[str] = Query(None),
    severity: Optional[str] = Query(None),
    risk_band: Optional[str] = Query(None),
    inactivity_status: Optional[str] = Query(None),
):
    """KPIs: drivers_monitored, sharp_degradation, sustained_degradation, recovery_cases, dormant_risk_cases, high_value_at_risk, avg_days_since_last_trip."""
    try:
        return await _run_sync(
            get_driver_behavior_summary,
            recent_weeks=recent_weeks,
            baseline_weeks=baseline_weeks,
            as_of_week=as_of_week,
            country=country,
            city=city,
            park_id=park_id,
            segment_current=segment_current,
            alert_type=alert_type,
            severity=severity,
            risk_band=risk_band,
            inactivity_status=inactivity_status,
        )
    except Exception as e:
        logger.error("GET /ops/driver-behavior/summary: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/driver-behavior/drivers")
async def get_driver_behavior_drivers_endpoint(
    recent_weeks: int = Query(4, ge=1, le=32),
    baseline_weeks: int = Query(16, ge=1, le=32),
    as_of_week: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    segment_current: Optional[str] = Query(None),
    alert_type: Optional[str] = Query(None),
    severity: Optional[str] = Query(None),
    risk_band: Optional[str] = Query(None),
    inactivity_status: Optional[str] = Query(None),
    min_baseline_trips: Optional[float] = Query(None),
    limit: int = Query(500, ge=1, le=2000),
    offset: int = Query(0, ge=0),
    order_by: str = Query("risk_score"),
    order_dir: str = Query("desc"),
):
    """Lista de conductores con desviación (ventanas reciente/baseline), days_since_last_trip, alert_type, risk_score."""
    try:
        return await _run_sync(
            get_driver_behavior_drivers,
            recent_weeks=recent_weeks,
            baseline_weeks=baseline_weeks,
            as_of_week=as_of_week,
            country=country,
            city=city,
            park_id=park_id,
            segment_current=segment_current,
            alert_type=alert_type,
            severity=severity,
            risk_band=risk_band,
            inactivity_status=inactivity_status,
            min_baseline_trips=min_baseline_trips,
            limit=limit,
            offset=offset,
            order_by=order_by,
            order_dir=order_dir,
        )
    except Exception as e:
        logger.error("GET /ops/driver-behavior/drivers: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/driver-behavior/driver-detail")
async def get_driver_behavior_driver_detail_endpoint(
    driver_key: str = Query(..., description="ID del conductor"),
    recent_weeks: int = Query(4, ge=1, le=32),
    baseline_weeks: int = Query(16, ge=1, le=32),
    as_of_week: Optional[str] = Query(None),
):
    """Detalle de un conductor: métricas de desviación + serie semanal para gráfico."""
    try:
        return await _run_sync(
            get_driver_behavior_driver_detail,
            driver_key=driver_key,
            recent_weeks=recent_weeks,
            baseline_weeks=baseline_weeks,
            as_of_week=as_of_week,
        )
    except Exception as e:
        logger.error("GET /ops/driver-behavior/driver-detail: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/driver-behavior/export")
async def get_driver_behavior_export_endpoint(
    recent_weeks: int = Query(4, ge=1, le=32),
    baseline_weeks: int = Query(16, ge=1, le=32),
    as_of_week: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    segment_current: Optional[str] = Query(None),
    alert_type: Optional[str] = Query(None),
    severity: Optional[str] = Query(None),
    risk_band: Optional[str] = Query(None),
    inactivity_status: Optional[str] = Query(None),
    max_rows: int = Query(10000, ge=1, le=50000),
):
    """Exporta conductores con filtros activos (CSV)."""
    try:
        rows = await _run_sync(
            get_driver_behavior_export,
            recent_weeks=recent_weeks,
            baseline_weeks=baseline_weeks,
            as_of_week=as_of_week,
            country=country,
            city=city,
            park_id=park_id,
            segment_current=segment_current,
            alert_type=alert_type,
            severity=severity,
            risk_band=risk_band,
            inactivity_status=inactivity_status,
            max_rows=max_rows,
        )
        rows = rows or []
        cols = ["driver_key", "driver_name", "country", "city", "park_name", "recent_window_weeks", "baseline_window_weeks", "recent_avg_weekly_trips", "baseline_avg_weekly_trips", "delta_pct", "behavior_direction", "days_since_last_trip", "alert_type", "risk_score", "risk_band", "suggested_action"]
        body = _to_csv(rows, cols)
        return Response(content=body, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=driver_behavior.csv"})
    except Exception as e:
        logger.error("GET /ops/driver-behavior/export: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# --- Top Driver Behavior (Elite/Legend benchmarks) ---
@router.get("/top-driver-behavior/summary")
async def get_top_driver_behavior_summary_endpoint(
    week_start: Optional[str] = Query(None),
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
):
    """Resumen: elite_drivers, legend_drivers, ft_drivers."""
    try:
        return await _run_sync(
            get_top_driver_behavior_summary,
            week_start=week_start, from_date=from_, to_date=to,
            country=country, city=city, park_id=park_id,
        )
    except Exception as e:
        logger.error("GET /ops/top-driver-behavior/summary: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/top-driver-behavior/benchmarks")
async def get_top_driver_behavior_benchmarks_endpoint(
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
):
    """Benchmarks por segment (ELITE, LEGEND, FT)."""
    try:
        return await _run_sync(
            get_top_driver_behavior_benchmarks,
            country=country, city=city, park_id=park_id,
        )
    except Exception as e:
        logger.error("GET /ops/top-driver-behavior/benchmarks: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/top-driver-behavior/patterns")
async def get_top_driver_behavior_patterns_endpoint(
    segment_current: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    limit: int = Query(200, ge=1, le=1000),
):
    """Concentración por segment, ciudad, parque."""
    try:
        return await _run_sync(
            get_top_driver_behavior_patterns,
            segment_current=segment_current, country=country, city=city, limit=limit,
        )
    except Exception as e:
        logger.error("GET /ops/top-driver-behavior/patterns: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/top-driver-behavior/playbook-insights")
async def get_top_driver_behavior_playbook_insights_endpoint(
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
):
    """Insights tipo playbook (Elite vs FT, etc.)."""
    try:
        insights = await _run_sync(get_top_driver_behavior_playbook_insights, country=country, city=city)
        return {"data": insights}
    except Exception as e:
        logger.error("GET /ops/top-driver-behavior/playbook-insights: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/top-driver-behavior/export")
async def get_top_driver_behavior_export_endpoint(
    segment_current: Optional[str] = Query(None),
    week_start: Optional[str] = Query(None),
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    format: Optional[str] = Query("csv", description="csv | excel"),
    max_rows: int = Query(10000, ge=1, le=50000),
):
    """Exporta lista Elite/Legend/FT con filtros."""
    try:
        rows = await _run_sync(
            get_top_driver_behavior_export,
            segment_current=segment_current, week_start=week_start, from_date=from_, to_date=to,
            country=country, city=city, park_id=park_id, max_rows=max_rows,
        )
        cols = ["driver_key", "driver_name", "week_start", "week_label", "country", "city", "park_id", "park_name", "segment_current", "trips_current_week", "avg_trips_baseline", "consistency_score", "active_weeks_in_window"]
        if (format or "csv").lower() == "excel":
            try:
                import openpyxl
                from io import BytesIO
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Top Driver Behavior"
                ws.append(cols)
                for r in rows:
                    ws.append([r.get(c) for c in cols])
                buf = BytesIO()
                wb.save(buf)
                return Response(
                    content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=top_driver_behavior.xlsx"},
                )
            except ImportError:
                pass
        body = _to_csv(rows, cols)
        return Response(content=body, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=top_driver_behavior.csv"})
    except Exception as e:
        logger.error("GET /ops/top-driver-behavior/export: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# --- Data Freshness & Coverage (fuentes base + derivados) ---
@router.get("/data-freshness")
async def get_data_freshness_endpoint(
    latest_only: bool = Query(True, description="Solo última ejecución por dataset"),
):
    """Auditoría de freshness por dataset: source_max_date, derived_max_date, expected_latest_date, status (OK, PARTIAL_EXPECTED, LAGGING, MISSING_EXPECTED_DATA)."""
    try:
        return get_freshness_audit(latest_only=latest_only)
    except Exception as e:
        logger.error("GET /ops/data-freshness: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/data-freshness/alerts")
async def get_data_freshness_alerts_endpoint():
    """Resumen de alertas accionables: datasets con status distinto de OK y mensaje explicativo."""
    try:
        return get_freshness_alerts()
    except Exception as e:
        logger.error("GET /ops/data-freshness/alerts: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/data-freshness/expectations")
async def get_data_freshness_expectations_endpoint():
    """Configuración de expectativas por dataset (grain, expected_delay_days, source/derived objects)."""
    try:
        return get_freshness_expectations()
    except Exception as e:
        logger.error("GET /ops/data-freshness/expectations: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/data-freshness/global")
async def get_data_freshness_global_endpoint(
    group: str | None = Query(None, description="operational = solo grupo REAL; legacy = solo legacy; omitir = todos"),
):
    """Estado global de frescura para el banner de UI. group=operational para pestaña REAL (no falla por datasets legacy)."""
    try:
        return get_freshness_global_status(group=group)
    except Exception as e:
        logger.error("GET /ops/data-freshness/global: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-margin-quality")
@router.get("/real/margin-quality")
async def get_real_margin_quality_endpoint(
    days_recent: int = Query(90, ge=1, le=365, description="Días atrás para agregado"),
    findings_limit: int = Query(20, ge=1, le=100, description="Límite de hallazgos persistidos"),
):
    """Estado de calidad de margen en fuente REAL. Alias: GET /ops/real-margin-quality y GET /ops/real/margin-quality."""
    # #region agent log
    _debug_log_ops("ops.py:margin-quality", "request_start", {"path": "/real-margin-quality"}, "H1")
    # #endregion
    t0 = _time.perf_counter()
    try:
        out = await _run_sync(get_margin_quality_full, days_recent=days_recent, findings_limit=findings_limit)
        # #region agent log
        _debug_log_ops("ops.py:margin-quality", "request_ok", {"path": "/real-margin-quality", "duration_ms": round((_time.perf_counter() - t0) * 1000)}, "H1")
        # #endregion
        return out
    except Exception as e:
        # #region agent log
        _debug_log_ops("ops.py:margin-quality", "request_err", {"path": "/real-margin-quality", "err": str(e), "duration_ms": round((_time.perf_counter() - t0) * 1000)}, "H1")
        # #endregion
        logger.error("GET /ops/real/margin-quality: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/real/margin-quality/run")
async def post_real_margin_quality_run():
    """Ejecuta auditoría de huecos de margen (audit_real_margin_source_gaps) y persiste en ops.real_margin_quality_audit. Uso: cron o admin."""
    try:
        import subprocess
        backend_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        r = subprocess.run(
            [sys.executable, "-m", "scripts.audit_real_margin_source_gaps", "--days", "90", "--persist"],
            cwd=backend_dir,
            capture_output=True,
            text=True,
            timeout=180,
        )
        return {
            "ok": r.returncode == 0,
            "stdout": r.stdout or "",
            "stderr": r.stderr or "",
            "returncode": r.returncode,
        }
    except subprocess.TimeoutExpired:
        logger.error("POST /ops/real/margin-quality/run: timeout")
        raise HTTPException(status_code=504, detail="Audit run timeout")
    except Exception as e:
        logger.error("POST /ops/real/margin-quality/run: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/data-trust")
async def get_data_trust_endpoint(
    view: str = Query(..., description="Vista: plan_vs_real | real_lob | driver_lifecycle | supply | resumen"),
):
    """
    Capa Data Trust: estado de confianza de data por vista.
    Retorna { data_trust: { status: ok|warning|blocked, message, last_update } }.
    Si falla el cálculo → status warning, message "Estado de data no disponible".
    """
    try:
        from app.services.data_trust_service import get_data_trust_status
        data_trust = get_data_trust_status(view_name=view.strip().lower())
        return {"data_trust": data_trust}
    except Exception as e:
        logger.debug("GET /ops/data-trust: %s", e)
        return {
            "data_trust": {
                "status": "warning",
                "message": "Estado de data no disponible",
                "last_update": None,
            }
        }


@router.get("/data-confidence")
async def get_data_confidence_endpoint(
    view: str = Query(..., description="Vista: real_lob | resumen | plan_vs_real | supply | driver_lifecycle | real_vs_projection | behavioral_alerts | leakage | real_margin_quality"),
):
    """
    Observabilidad: detalle del Confidence Engine por vista.
    Retorna source_of_truth, source_mode, freshness_status, completeness_status, consistency_status,
    confidence_score, trust_status, message, last_update, details.
    """
    try:
        from app.services.confidence_engine import get_confidence_status
        return get_confidence_status(view.strip().lower(), None)
    except Exception as e:
        logger.debug("GET /ops/data-confidence: %s", e)
        from app.services.confidence_engine import _fallback_response
        return _fallback_response(
            view.strip().lower(),
            "Estado de data no disponible",
            source_of_truth=None,
            source_mode="unknown",
        )


@router.get("/data-confidence/registry")
async def get_data_confidence_registry_endpoint():
    """
    Observabilidad: registro Source of Truth. Qué fuente manda hoy por dominio.
    """
    try:
        from app.config.source_of_truth_registry import SOURCE_OF_TRUTH, REGISTERED_VIEWS
        return {
            "registry": SOURCE_OF_TRUTH,
            "registered_views": list(REGISTERED_VIEWS),
        }
    except Exception as e:
        logger.error("GET /ops/data-confidence/registry: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/data-confidence/summary")
async def get_data_confidence_summary_endpoint():
    """
    Observabilidad: resumen de confianza de todas las vistas registradas.
    """
    try:
        from app.services.confidence_engine import get_confidence_summary
        return get_confidence_summary()
    except Exception as e:
        logger.error("GET /ops/data-confidence/summary: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/decision-signal")
async def get_decision_signal_endpoint(
    view: str = Query(..., description="Vista: real_lob | resumen | plan_vs_real | supply | driver_lifecycle | ..."),
):
    """
    Decision Layer: señal operativa por vista (action, priority, message, reason).
    Fuente única: Confidence Engine + view_criticality.
    """
    try:
        from app.services.decision_engine import get_decision_signal
        return get_decision_signal(view.strip().lower(), None)
    except Exception as e:
        logger.debug("GET /ops/decision-signal: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/decision-signal/summary")
async def get_decision_signal_summary_endpoint():
    """
    Resumen de decisiones por vista: view, action, priority.
    """
    try:
        from app.services.decision_engine import get_decision_signal_summary
        return get_decision_signal_summary()
    except Exception as e:
        logger.error("GET /ops/decision-signal/summary: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-source-status")
async def get_real_source_status_endpoint():
    """
    Gobierno: estado de fuente REAL por pantalla/feature.
    source_status: canonical | legacy | migrating | source_incomplete.
    Resumen usa canónica (mv_real_monthly_canonical_hist) desde cierre Fase 1.
    """
    try:
        screens = [
            {
                "screen_id": "performance_resumen",
                "screen_name": "Performance > Resumen",
                "source_status": "canonical",
                "message": "Fuente canónica (mv_real_monthly_canonical_hist)",
            },
            {
                "screen_id": "performance_real_daily",
                "screen_name": "Performance > Real (diario)",
                "source_status": "canonical",
                "message": "Fuente canónica",
            },
            {
                "screen_id": "operacion_drill",
                "screen_name": "Operación > Drill",
                "source_status": "canonical",
                "message": "Fuente canónica",
            },
            {
                "screen_id": "performance_plan_vs_real",
                "screen_name": "Performance > Plan vs Real",
                "source_status": "migrating",
                "message": "Migrando a fuente canónica",
            },
            {
                "screen_id": "proyeccion_real_vs_projection",
                "screen_name": "Proyección > Real vs Proyección",
                "source_status": "source_incomplete",
                "message": "Vista temporalmente limitada; puede depender de objetos faltantes",
            },
            {
                "screen_id": "risk_behavioral_alerts",
                "screen_name": "Riesgo > Alertas de conducta",
                "source_status": "under_review",
                "message": "En revisión; tiempos de respuesta pueden ser elevados",
            },
            {
                "screen_id": "risk_fleet_leakage",
                "screen_name": "Riesgo > Fuga de flota",
                "source_status": "under_review",
                "message": "En revisión; validar estabilidad en runtime",
            },
        ]
        return {
            "screens": screens,
            "resumen_uses_canonical": True,
        }
    except Exception as e:
        logger.error("GET /ops/real-source-status: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/data-pipeline-health")
async def get_data_pipeline_health_endpoint(
    latest_only: bool = Query(True, description="Solo última ejecución por dataset"),
):
    """Centro de observabilidad del pipeline: por dataset source_max_date, derived_max_date, lag_days, expected_latest_date, status, alert_reason, last_checked_at."""
    try:
        audit = get_freshness_audit(latest_only=latest_only)
        return {
            "datasets": audit,
            "last_checked": audit[0].get("checked_at") if audit else None,
        }
    except Exception as e:
        logger.error("GET /ops/data-pipeline-health: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/integrity-report")
async def get_integrity_report_endpoint():
    """Reporte global de integridad: check_name, status, severity, details (TRIP LOSS, B2B, LOB MAPPING, DUPLICATES, MV STALE, JOIN LOSS, WEEKLY ANOMALY)."""
    try:
        return get_integrity_report()
    except Exception as e:
        logger.error("GET /ops/integrity-report: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/system-health")
async def get_system_health_endpoint():
    """Estado del sistema para dashboard System Health: integridad, freshness MVs, ingestión, última auditoría."""
    try:
        return get_system_health()
    except Exception as e:
        logger.error("GET /ops/system-health: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/diagnostics/serving-sources")
async def get_serving_sources_diagnostics():
    """Serving enforcement diagnostics: compliance, freshness, forbidden sources, hard-gate status."""
    try:
        from app.utils.source_trace import get_serving_diagnostics_full, get_unguarded_features
        from app.services.serving_guardrails import (
            get_usage_log, FORBIDDEN_SERVING_SOURCES, get_all_declared_policies,
            get_db_guard_mode, get_db_gate_log,
        )
        from app.db.connection import get_db

        with get_db() as conn:
            diagnostics = get_serving_diagnostics_full(conn)

        compliant = sum(1 for d in diagnostics if d["compliance_status"] == "COMPLIANT")
        warnings = sum(1 for d in diagnostics if d["compliance_status"] == "WARNING")
        non_compliant = sum(1 for d in diagnostics if d["compliance_status"] == "NON_COMPLIANT")
        unknown = sum(1 for d in diagnostics if d["compliance_status"] == "UNKNOWN")

        policies = get_all_declared_policies()
        unguarded = get_unguarded_features()

        return {
            "diagnostics": diagnostics,
            "total": len(diagnostics),
            "summary": {
                "compliant": compliant,
                "warnings": warnings,
                "non_compliant": non_compliant,
                "unknown": unknown,
                "policies_declared": len(policies),
                "unguarded_features": len(unguarded),
                "db_guard_mode": get_db_guard_mode().value,
            },
            "forbidden_serving_sources": FORBIDDEN_SERVING_SOURCES,
            "unguarded_features": unguarded,
            "recent_usage_log": get_usage_log()[-20:],
            "db_gate_log": get_db_gate_log()[-10:],
        }
    except Exception as e:
        logger.error("GET /ops/diagnostics/serving-sources: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/integrity-audit/run")
async def post_integrity_audit_run():
    """Ejecuta auditoría de integridad (audit_control_tower.py) y persiste en ops.data_integrity_audit. Uso: cron o admin."""
    try:
        import subprocess
        backend_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        r = subprocess.run(
            [sys.executable, "-m", "scripts.audit_control_tower"],
            cwd=backend_dir,
            capture_output=True,
            text=True,
            timeout=120,
        )
        return {
            "ok": r.returncode == 0,
            "stdout": r.stdout or "",
            "stderr": r.stderr or "",
            "returncode": r.returncode,
        }
    except subprocess.TimeoutExpired:
        logger.error("POST /ops/integrity-audit/run: timeout")
        raise HTTPException(status_code=504, detail="Audit run timeout")
    except Exception as e:
        logger.error("POST /ops/integrity-audit/run: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/data-freshness/run")
async def post_data_freshness_run():
    """Ejecuta el chequeo de freshness y escribe en ops.data_freshness_audit. Uso: cron o admin."""
    try:
        import subprocess
        # __file__ = backend/app/routers/ops.py -> backend = dirname(dirname(dirname(__file__)))
        backend_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        r = subprocess.run(
            [sys.executable, "-m", "scripts.run_data_freshness_audit"],
            cwd=backend_dir,
            capture_output=True,
            text=True,
            timeout=300,
        )
        return {
            "ok": r.returncode == 0,
            "stdout": r.stdout or "",
            "stderr": r.stderr or "",
            "returncode": r.returncode,
        }
    except subprocess.TimeoutExpired:
        logger.error("POST /ops/data-freshness/run: timeout")
        raise HTTPException(status_code=504, detail="Audit run timeout")
    except Exception as e:
        logger.error("POST /ops/data-freshness/run: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/pipeline-refresh")
async def post_pipeline_refresh(
    skip_backfill: bool = Query(False, description="Omitir backfill Real LOB"),
    skip_driver: bool = Query(False, description="Omitir refresh driver lifecycle"),
    skip_supply: bool = Query(False, description="Omitir refresh supply"),
    skip_audit: bool = Query(False, description="Omitir auditoría final"),
):
    """Ejecuta pipeline unificado: backfill Real LOB (mes actual+anterior), refresh driver lifecycle, refresh supply, auditoría. Uso: cron o admin. Puede tardar varios minutos."""
    try:
        backend_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        cmd = [sys.executable, "-m", "scripts.run_pipeline_refresh_and_audit"]
        if skip_backfill:
            cmd.append("--skip-backfill")
        if skip_driver:
            cmd.append("--skip-driver")
        if skip_supply:
            cmd.append("--skip-supply")
        if skip_audit:
            cmd.append("--skip-audit")
        r = subprocess.run(
            cmd,
            cwd=backend_dir,
            capture_output=True,
            text=True,
            timeout=3600,
        )
        return {
            "ok": r.returncode == 0,
            "stdout": r.stdout or "",
            "stderr": r.stderr or "",
            "returncode": r.returncode,
        }
    except subprocess.TimeoutExpired:
        logger.error("POST /ops/pipeline-refresh: timeout")
        raise HTTPException(status_code=504, detail="Pipeline refresh timeout")
    except Exception as e:
        logger.error("POST /ops/pipeline-refresh: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/supply/refresh")
async def post_supply_refresh():
    """Refresca MVs de Supply Alerting. CONCURRENTLY."""
    try:
        refresh_supply_alerting_mvs()
        return {"ok": True, "message": "ops.refresh_supply_alerting_mvs() ejecutado"}
    except Exception as e:
        logger.error("POST /ops/supply/refresh: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/supply/refresh-alerting")
async def post_supply_refresh_alerting():
    """Refresca MVs de Supply Alerting (solo si SUPPLY_REFRESH_ALLOWED=true). Uso admin."""
    import os
    if os.environ.get("SUPPLY_REFRESH_ALLOWED", "").lower() not in ("1", "true", "yes"):
        raise HTTPException(status_code=403, detail="Supply refresh not allowed (set SUPPLY_REFRESH_ALLOWED)")
    try:
        refresh_supply_alerting_mvs()
        return {"ok": True, "message": "ops.refresh_supply_alerting_mvs() ejecutado"}
    except Exception as e:
        logger.error("POST /ops/supply/refresh-alerting: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-lob/filters")
async def get_real_lob_filters_endpoint(
    country: Optional[str] = Query(None, description="Filtrar ciudades/parks por país"),
    city: Optional[str] = Query(None, description="Filtrar parks por ciudad"),
):
    """Opciones para dropdowns: countries, cities, parks, lob_groups, tipo_servicio, segments, years. Cache 5 min."""
    try:
        return get_real_lob_filters(country=country, city=city)
    except Exception as e:
        logger.error(f"Error Real LOB filters: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-lob/v2/data")
async def get_real_lob_v2_data_endpoint(
    period_type: Literal["monthly", "weekly"] = Query("monthly"),
    agg_level: str = Query("DETALLE", description="DETALLE|TOTAL_PAIS|TOTAL_CIUDAD|TOTAL_PARK|PARK_X_MES|PARK_X_MES_X_LOB|PARK_X_SEMANA|PARK_X_SEMANA_X_LOB"),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
    lob_group: Optional[str] = Query(None),
    tipo_servicio: Optional[str] = Query(None),
    segment_tag: Optional[str] = Query(None, description="Todos|B2B|B2C"),
    year: Optional[int] = Query(None, description="Año; si vacío, últimos 12 meses"),
):
    """Datos Real LOB v2 con consolidación. Devuelve totals (trips, b2b_ratio, rows), rows y meta."""
    try:
        return get_real_lob_v2_data(
            period_type=period_type,
            agg_level=agg_level,
            country=country,
            city=city,
            park_id=park_id,
            lob_group=lob_group,
            tipo_servicio=tipo_servicio,
            segment_tag=segment_tag,
            year=year,
        )
    except Exception as e:
        logger.error(f"Error Real LOB v2 data: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ─── Real Operational (hourly-first: today, yesterday, week, day view, hourly view, cancellations, comparatives) ─────────────────
@router.get("/real-operational/snapshot")
async def get_real_operational_snapshot_endpoint(
    window: Literal["today", "yesterday", "this_week"] = Query("today"),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    park_id: Optional[str] = Query(None),
):
    """Snapshot operativo: hoy, ayer o esta semana (pedidos, completados, cancelados, revenue, margin, duración, tasas)."""
    try:
        return await _run_sync(
            get_operational_snapshot,
            window=window,
            country=country,
            city=city,
            park_id=park_id,
        )
    except Exception as e:
        logger.error("GET /ops/real-operational/snapshot: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-operational/day-view")
async def get_real_operational_day_view_endpoint(
    days_back: int = Query(14, ge=1, le=365),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    group_by: Literal["day", "city", "park", "lob", "service"] = Query("day"),
):
    """Vista por día (últimos N días) con drill por día/ciudad/park/lob/servicio."""
    try:
        return await _run_sync(
            get_day_view,
            days_back=days_back,
            country=country,
            city=city,
            group_by=group_by,
        )
    except Exception as e:
        logger.error("GET /ops/real-operational/day-view: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-operational/hourly-view")
async def get_real_operational_hourly_view_endpoint(
    days_back: int = Query(7, ge=1, le=90),
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    group_by: Literal["hour", "city", "park", "lob", "service"] = Query("hour"),
):
    """Vista por hora del día (0-23) con drill opcional por ciudad/park/lob/servicio."""
    try:
        return await _run_sync(
            get_hourly_view,
            days_back=days_back,
            country=country,
            city=city,
            group_by=group_by,
        )
    except Exception as e:
        logger.error("GET /ops/real-operational/hourly-view: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-operational/cancellations")
async def get_real_operational_cancellations_endpoint(
    days_back: int = Query(14, ge=1, le=90),
    country: Optional[str] = Query(None),
    limit: int = Query(20, ge=5, le=100),
    by: Literal["reason", "reason_group", "hour", "city", "park", "service"] = Query("reason_group"),
):
    """Top motivos de cancelación; agrupado por razón, hora, ciudad, park o servicio."""
    try:
        return await _run_sync(
            get_cancellation_view,
            days_back=days_back,
            country=country,
            limit=limit,
            by=by,
        )
    except Exception as e:
        logger.error("GET /ops/real-operational/cancellations: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-operational/comparatives/today-vs-yesterday")
async def get_real_operational_today_vs_yesterday_endpoint(
    country: Optional[str] = Query(None),
):
    """Comparativo: hoy vs ayer (pedidos, completados, cancelados, revenue, margin, cancel rate, duración)."""
    try:
        return await _run_sync(get_today_vs_yesterday, country=country)
    except Exception as e:
        logger.error("GET /ops/real-operational/comparatives/today-vs-yesterday: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-operational/comparatives/today-vs-same-weekday")
async def get_real_operational_today_vs_same_weekday_endpoint(
    n_weeks: int = Query(4, ge=1, le=12),
    country: Optional[str] = Query(None),
):
    """Comparativo: hoy vs promedio de los últimos N mismos días de semana (ej. últimos 4 lunes)."""
    try:
        return await _run_sync(get_today_vs_same_weekday_avg, n_weeks=n_weeks, country=country)
    except Exception as e:
        logger.error("GET /ops/real-operational/comparatives/today-vs-same-weekday: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-operational/comparatives/current-hour-vs-historical")
async def get_real_operational_current_hour_vs_historical_endpoint(
    country: Optional[str] = Query(None),
    weeks_back: int = Query(4, ge=1, le=12),
):
    """Comparativo: hora actual vs mismo tramo horario en semanas anteriores."""
    try:
        return await _run_sync(get_current_hour_vs_historical, country=country, weeks_back=weeks_back)
    except Exception as e:
        logger.error("GET /ops/real-operational/comparatives/current-hour-vs-historical: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-operational/comparatives/this-week-vs-comparable")
async def get_real_operational_this_week_vs_comparable_endpoint(
    country: Optional[str] = Query(None),
    weeks_back: int = Query(4, ge=1, le=12),
):
    """Comparativo: esta semana (lunes a hoy) vs promedio de las últimas N semanas."""
    try:
        return await _run_sync(get_this_week_vs_comparable, country=country, weeks_back=weeks_back)
    except Exception as e:
        logger.error("GET /ops/real-operational/comparatives/this-week-vs-comparable: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# ─── Real LOB Drill PRO (rutas más específicas primero para evitar 404) [legacy: preferir real-operational] ─────────────────
@router.get("/real-lob/drill/parks")
async def get_real_lob_drill_parks_endpoint(
    country: Optional[str] = Query(None, description="Filtrar parks por país (pe | co)"),
):
    """Lista de parks para el filtro Park del drill."""
    try:
        parks = await _run_sync(get_real_lob_drill_parks, country=country)
        return {"parks": parks}
    except Exception as e:
        logger.error("GET /ops/real-lob/drill/parks: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-lob/drill/children")
async def get_real_lob_drill_children_endpoint(
    country: str = Query(..., description="País (pe | co)"),
    period: Literal["month", "week"] = Query("month"),
    period_start: str = Query(..., description="YYYY-MM-DD o YYYY-MM-01"),
    desglose: Literal["LOB", "PARK", "SERVICE_TYPE"] = Query("PARK"),
    segmento: Optional[Literal["all", "b2c", "b2b"]] = Query("all"),
    drill_lob_id: Optional[str] = Query(None, description="Filtro LOB (solo válido si desglose=LOB)"),
    drill_park_id: Optional[str] = Query(None, description="Filtro Park (válido si desglose=PARK o desglose=SERVICE_TYPE para filtrar tipo_servicio por park)"),
    park_id: Optional[str] = Query(None, description="Filtro por park (igual que drill_park_id; aplica a desglose tipo_servicio)"),
):
    """
    Desglose por LOB (1 fila por lob_group), Park (city, park_name), o Tipo de servicio. Orden: viajes DESC.
    Si desglose=SERVICE_TYPE y park_id (o drill_park_id) está indicado, el desglose se limita a ese park.

    Contrato params por dimensión:
    - desglose=LOB  => permitido drill_lob_id; 400 si llega drill_park_id.
    - desglose=PARK => permitido drill_park_id; 400 si llega drill_lob_id.
    - desglose=SERVICE_TYPE => permitido park_id/drill_park_id para filtrar por park.
    """
    if desglose == "PARK" and drill_lob_id is not None and str(drill_lob_id).strip() != "":
        raise HTTPException(
            status_code=400,
            detail="Incompatible drill params for groupBy=PARK: drill_lob_id is not allowed when desglose is PARK.",
        )
    if desglose == "LOB" and (drill_park_id is not None and str(drill_park_id).strip() != "" or park_id is not None and str(park_id).strip() != ""):
        raise HTTPException(
            status_code=400,
            detail="Incompatible drill params for groupBy=LOB: drill_park_id/park_id is not allowed when desglose is LOB.",
        )
    effective_park_id = (park_id or drill_park_id)
    if effective_park_id is not None:
        effective_park_id = str(effective_park_id).strip() or None
    # #region agent log
    _debug_log_ops("ops.py:children", "request_start", {"path": "/real-lob/drill/children", "desglose": desglose, "period": period}, "H4")
    # #endregion
    t0_children = _time.perf_counter()
    try:
        data = await _run_sync(
            get_real_lob_drill_pro_children,
            country=country,
            period=period,
            period_start=period_start,
            desglose=desglose,
            segmento=segmento,
            park_id=effective_park_id,
        )
        # #region agent log
        _debug_log_ops("ops.py:children", "request_ok", {"path": "/real-lob/drill/children", "duration_ms": round((_time.perf_counter() - t0_children) * 1000)}, "H4")
        # #endregion
        return {"data": data}
    except Exception as e:
        # #region agent log
        _debug_log_ops("ops.py:children", "request_err", {"path": "/real-lob/drill/children", "err": str(e), "duration_ms": round((_time.perf_counter() - t0_children) * 1000)}, "H4")
        # #endregion
        logger.error("Real LOB drill PRO children: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-lob/drill")
async def get_real_lob_drill_pro_endpoint(
    period: Literal["month", "week"] = Query("month", description="month | week"),
    desglose: Literal["LOB", "PARK", "SERVICE_TYPE"] = Query("PARK", description="Desglose al expandir: LOB | PARK | SERVICE_TYPE"),
    segmento: Optional[Literal["all", "b2c", "b2b"]] = Query("all", description="all | b2c | b2b"),
    country: Optional[Literal["all", "pe", "co"]] = Query("all", description="all | pe | co"),
    park_id: Optional[str] = Query(None, description="Filtro opcional por park; aplica a timeline y a desglose tipo_servicio"),
):
    """Real LOB Drill PRO: countries[] con coverage, kpis, rows por periodo."""
    # #region agent log
    _debug_log_ops("ops.py:drill", "request_start", {"path": "/real-lob/drill", "period": period, "desglose": desglose}, "H2")
    # #endregion
    logger.info("Real LOB drill: request received period=%s desglose=%s segmento=%s park_id=%s", period, desglose, segmento, park_id)
    t0 = _time.perf_counter()
    try:
        out = await _run_sync(
            get_real_lob_drill_pro,
            period=period,
            desglose=desglose,
            segmento=segmento,
            country=country,
            park_id=park_id,
        )
        # #region agent log
        _debug_log_ops("ops.py:drill", "request_ok", {"path": "/real-lob/drill", "duration_ms": round((_time.perf_counter() - t0) * 1000)}, "H2")
        # #endregion
        return out
    except Exception as e:
        # #region agent log
        _debug_log_ops("ops.py:drill", "request_err", {"path": "/real-lob/drill", "err": str(e), "duration_ms": round((_time.perf_counter() - t0) * 1000)}, "H2")
        # #endregion
        logger.error("Real LOB drill PRO: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# ─── Period Semantics & Comparatives ─────────────────
@router.get("/period-semantics")
async def get_period_semantics_endpoint(
    reference: Optional[str] = Query(None, description="Fecha referencia YYYY-MM-DD (default: hoy)"),
):
    """Semántica temporal: last_closed_day/week/month, current_open_week/month y labels para UI."""
    try:
        ref = None
        if reference and len(reference.strip()) >= 10:
            from datetime import date
            ref = date.fromisoformat(reference.strip()[:10])
        return await _run_sync(get_period_semantics, ref)
    except Exception as e:
        logger.error("GET /ops/period-semantics: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-lob/comparatives/weekly")
async def get_real_lob_wow_endpoint(
    country: Optional[str] = Query(None, description="Filtrar por país (pe | co)"),
):
    """WoW: última semana cerrada vs semana cerrada anterior. Métricas: viajes, margen_total, margen_trip, km_prom, b2b_pct."""
    try:
        return await _run_sync(get_weekly_comparative, country=country)
    except Exception as e:
        logger.error("GET /ops/real-lob/comparatives/weekly: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-lob/comparatives/monthly")
async def get_real_lob_mom_endpoint(
    country: Optional[str] = Query(None, description="Filtrar por país (pe | co)"),
):
    """MoM: último mes cerrado vs mes cerrado anterior. Métricas: viajes, margen_total, margen_trip, km_prom, b2b_pct."""
    try:
        return await _run_sync(get_monthly_comparative, country=country)
    except Exception as e:
        logger.error("GET /ops/real-lob/comparatives/monthly: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-lob/daily/summary")
async def get_real_lob_daily_summary_endpoint(
    day: Optional[str] = Query(None, description="Fecha YYYY-MM-DD (default: último día cerrado = ayer)"),
    country: Optional[str] = Query(None, description="Filtrar por país (pe | co)"),
):
    """Vista diaria: KPIs agregados por día (viajes, margen, km_prom, B2B %)."""
    try:
        return await _run_sync(get_daily_summary, day=day, country=country)
    except Exception as e:
        logger.error("GET /ops/real-lob/daily/summary: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-lob/daily/comparative")
async def get_real_lob_daily_comparative_endpoint(
    day: Optional[str] = Query(None, description="Fecha YYYY-MM-DD (default: último día cerrado)"),
    country: Optional[str] = Query(None, description="Filtrar por país (pe | co)"),
    baseline: Literal["D-1", "same_weekday_previous_week", "same_weekday_avg_4w"] = Query(
        "D-1",
        description="D-1 = vs día anterior; same_weekday_previous_week = vs mismo día semana pasada; same_weekday_avg_4w = vs promedio 4 mismos días",
    ),
):
    """Comparativo diario: día consultado vs baseline (D-1, mismo día semana pasada, o promedio 4 mismos días)."""
    try:
        return await _run_sync(get_daily_comparative, day=day, country=country, baseline=baseline)
    except Exception as e:
        logger.error("GET /ops/real-lob/daily/comparative: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-lob/daily/table")
async def get_real_lob_daily_table_endpoint(
    day: Optional[str] = Query(None, description="Fecha YYYY-MM-DD (default: último día cerrado)"),
    country: Optional[str] = Query(None, description="Filtrar por país (pe | co)"),
    group_by: Literal["lob", "park"] = Query("lob", description="Agrupar por LOB o por Park"),
    baseline: Optional[Literal["D-1", "same_weekday_previous_week", "same_weekday_avg_4w"]] = Query(
        None, description="Si se indica, cada fila incluye *_baseline y *_delta_pct (comparativo por fila)",
    ),
):
    """Tabla diaria: filas por LOB o por Park. Con baseline se añaden columnas comparativas por fila."""
    try:
        return await _run_sync(get_daily_table, day=day, country=country, group_by=group_by, baseline=baseline)
    except Exception as e:
        logger.error("GET /ops/real-lob/daily/table: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# ─── Real LOB Drill-down (timeline por país, drill LOB/Park) [legacy; preferir /real-lob/drill] ─────────────────
@router.get("/real-drill/summary")
async def get_real_drill_summary_endpoint(
    period_type: Literal["monthly", "weekly"] = Query("monthly", description="monthly | weekly"),
    segment: Optional[Literal["Todos", "B2B", "B2C"]] = Query("Todos", description="Todos | B2B | B2C"),
    limit_periods: Optional[int] = Query(None, description="Máx periodos (default 24 meses o 26 semanas)"),
):
    """
    Timeline por país: countries[] con coverage, kpis (sobre lo visible) y rows.
    Orden: PE primero, CO segundo. KPIs calculados sobre los periodos devueltos.
    """
    try:
        seg = None if segment == "Todos" else segment
        result = get_real_drill_summary_countries(
            period_type=period_type,
            segment=seg,
            limit_periods=limit_periods,
        )
        return result
    except RealDrillMvNotPopulatedError as e:
        return {
            "countries": [
                {"country": "pe", "coverage": {}, "kpis": {}, "rows": []},
                {"country": "co", "coverage": {}, "kpis": {}, "rows": []},
            ],
            "meta": {
                "last_period_monthly": None,
                "last_period_weekly": None,
                "hint": e.hint,
            },
        }
    except Exception as e:
        logger.error("Real drill summary: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-drill/by-lob")
async def get_real_drill_by_lob_endpoint(
    period_type: Literal["monthly", "weekly"] = Query("monthly"),
    country: str = Query(..., description="País (requerido)"),
    period_start: str = Query(..., description="Fecha inicio periodo (YYYY-MM-DD o YYYY-MM)"),
    segment: Optional[Literal["Todos", "B2B", "B2C"]] = Query("Todos"),
):
    """Desglose por LOB para un país y periodo. Orden: trips DESC."""
    try:
        seg = None if segment == "Todos" else segment
        data = get_real_drill_by_lob(
            period_type=period_type,
            country=country,
            period_start=period_start,
            segment=seg,
        )
        meta = get_real_drill_meta()
        return {
            "data": data,
            "meta": {
                "last_period_monthly": meta.get("last_period_monthly"),
                "last_period_weekly": meta.get("last_period_weekly"),
            },
        }
    except RealDrillMvNotPopulatedError as e:
        return {
            "data": [],
            "meta": {"last_period_monthly": None, "last_period_weekly": None, "hint": e.hint},
        }
    except Exception as e:
        logger.error("Real drill by-lob: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-drill/by-park")
async def get_real_drill_by_park_endpoint(
    period_type: Literal["monthly", "weekly"] = Query("monthly"),
    country: str = Query(..., description="País (requerido)"),
    period_start: str = Query(..., description="Fecha inicio periodo (YYYY-MM-DD)"),
    segment: Optional[Literal["Todos", "B2B", "B2C"]] = Query("Todos"),
):
    """Desglose por park para un país y periodo. Orden: trips DESC."""
    try:
        seg = None if segment == "Todos" else segment
        data = get_real_drill_by_park(
            period_type=period_type,
            country=country,
            period_start=period_start,
            segment=seg,
        )
        meta = get_real_drill_meta()
        return {
            "data": data,
            "meta": {
                "last_period_monthly": meta.get("last_period_monthly"),
                "last_period_weekly": meta.get("last_period_weekly"),
            },
        }
    except RealDrillMvNotPopulatedError as e:
        return {
            "data": [],
            "meta": {"last_period_monthly": None, "last_period_weekly": None, "hint": e.hint},
        }
    except Exception as e:
        logger.error("Real drill by-park: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/real-drill/refresh")
async def refresh_real_drill_endpoint():
    """
    Refresca la MV ops.mv_real_rollup_day. Uso interno (cron, admin).
    Ejecuta REFRESH MATERIALIZED VIEW CONCURRENTLY.
    """
    try:
        return refresh_real_drill_mv()
    except Exception as e:
        logger.error("Real drill refresh: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-drill/coverage")
async def get_real_drill_coverage_endpoint():
    """Cobertura por país: last_trip_date, last_month_with_data, last_week_with_data."""
    try:
        data = get_real_drill_coverage()
        return {"data": data}
    except Exception as e:
        msg = str(e) or ""
        if "does not exist" in msg or "relation" in msg.lower() or "aborted" in msg.lower():
            return {"data": []}
        logger.error("Real drill coverage: %s", e)
        raise HTTPException(status_code=500, detail=msg)


@router.get("/real-drill/totals")
async def get_real_drill_totals_endpoint(
    period_type: Literal["monthly", "weekly"] = Query("monthly"),
    segment: Optional[Literal["Todos", "B2B", "B2C"]] = Query("Todos"),
    limit_periods: Optional[int] = Query(None),
):
    """Totales (total_trips, b2b_ratio) sobre el rango mostrado en summary."""
    try:
        seg = None if segment == "Todos" else segment
        data = get_real_drill_totals(
            period_type=period_type,
            segment=seg,
            limit_periods=limit_periods,
        )
        return data
    except RealDrillMvNotPopulatedError as e:
        return {
            "total_trips": 0,
            "total_b2b_trips": 0,
            "b2b_ratio_pct": None,
            "margin_total": None,
            "margin_unit_avg_global": None,
            "distance_total_km": None,
            "distance_km_avg_global": None,
            "last_trip_ts": None,
            "hint": e.hint,
        }
    except Exception as e:
        logger.error("Real drill totals: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# ─── Real LOB Strategy (KPIs ejecutivos, forecast, rankings) ─────────────────
@router.get("/real-strategy/country")
async def get_real_strategy_country_endpoint(
    country: str = Query(..., description="País (requerido)"),
    year_real: Optional[int] = Query(None, description="Año para filtrar (opcional)"),
    segment_tag: Optional[Literal["B2B", "B2C"]] = Query(None, description="Segmento B2B/B2C (opcional)"),
    period_type: str = Query("monthly", description="Tipo de periodo (monthly por defecto)"),
):
    """
    KPIs estratégicos por país: total_trips_ytd, growth_mom, b2b_ratio, forecast_next_month,
    acceleration_index, concentration_index. Incluye tendencia 12 meses y ranking ciudades.
    """
    try:
        return get_real_strategy_country(
            country=country,
            year_real=year_real,
            segment_tag=segment_tag,
            period_type=period_type,
        )
    except Exception as e:
        logger.error(f"Error Real strategy country: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real-strategy/lob")
async def get_real_strategy_lob_endpoint(
    country: str = Query(..., description="País (requerido)"),
    year_real: Optional[int] = Query(None, description="Año para filtrar (opcional)"),
    segment_tag: Optional[Literal["B2B", "B2C"]] = Query(None, description="Segmento B2B/B2C (opcional)"),
    lob_group: Optional[str] = Query(None, description="Filtrar por LOB_GROUP (opcional)"),
    period_type: str = Query("monthly", description="Tipo de periodo (monthly por defecto)"),
):
    """
    Distribución LOB por país: trips, participation_pct, growth_mom, forecast_next_month, momentum_score.
    """
    try:
        return get_real_strategy_lob(
            country=country,
            year_real=year_real,
            segment_tag=segment_tag,
            lob_group=lob_group,
            period_type=period_type,
        )
    except Exception as e:
        logger.error(f"Error Real strategy LOB: {e}")
        raise HTTPException(status_code=500, detail=str(e))
@router.get("/real-strategy/cities")
async def get_real_strategy_cities_endpoint(
    country: str = Query(..., description="País (requerido)"),
    year_real: Optional[int] = Query(None, description="Año para filtrar (opcional)"),
    segment_tag: Optional[Literal["B2B", "B2C"]] = Query(None, description="Segmento B2B/B2C (opcional)"),
    period_type: str = Query("monthly", description="Tipo de periodo (monthly por defecto)"),
):
    """
    Ranking ciudades por país: city, trips, growth_mom, % país, expansion_index.
    """
    try:
        return get_real_strategy_cities(
            country=country,
            year_real=year_real,
            segment_tag=segment_tag,
            period_type=period_type,
        )
    except Exception as e:
        logger.error(f"Error Real strategy cities: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/real/monthly")
async def get_real_monthly_endpoint(
    country: Optional[str] = Query(None, description="Filtrar por país"),
    city: Optional[str] = Query(None, description="Filtrar por ciudad"),
    lob_base: Optional[str] = Query(None, description="Filtrar por línea de negocio base"),
    segment: Optional[str] = Query(None, description="Filtrar por segmento (b2b, b2c)"),
    year: int = Query(2025, description="Año del Real"),
    source: Optional[str] = Query(None, description="Si es 'canonical', usa cadena hourly-first (solo Resumen). Por defecto legacy."),
):
    """
    Obtiene datos REAL mensuales. Con source=canonical usa real_drill_dim_fact (cadena hourly-first).
    Sin source o source distinto de canonical usa ops.mv_real_trips_monthly (legacy).
    Retorna month, trips_real_completed, revenue_real_yego, active_drivers_real, avg_ticket_real.
    """
    try:
        if source and str(source).strip().lower() == "canonical":
            from app.services.canonical_real_monthly_service import get_real_monthly_canonical
            data = get_real_monthly_canonical(
                country=country,
                city=city,
                lob_base=lob_base,
                segment=segment,
                year=year,
            )
        else:
            data = get_real_monthly(
                country=country,
                city=city,
                lob_base=lob_base,
                segment=segment,
                year=year,
            )
        return {
            "data": data,
            "total_periods": len(data),
            "year": year,
            "source_used": "canonical" if (source and str(source).strip().lower() == "canonical") else "legacy",
        }
    except Exception as e:
        logger.error(f"Error al obtener Real monthly: {e}")
        raise HTTPException(status_code=500, detail=f"Error al obtener Real monthly: {str(e)}")

@router.get("/plan/monthly")
async def get_plan_monthly_endpoint(
    country: Optional[str] = Query(None, description="Filtrar por país"),
    city: Optional[str] = Query(None, description="Filtrar por ciudad"),
    lob_base: Optional[str] = Query(None, description="Filtrar por línea de negocio base"),
    segment: Optional[str] = Query(None, description="Filtrar por segmento (b2b, b2c)"),
    year: int = Query(2026, description="Año del Plan")
):
    """
    Obtiene datos PLAN mensuales agregados desde ops.v_plan_trips_monthly_latest.
    Retorna month, projected_trips, projected_revenue, projected_drivers, projected_ticket.
    """
    try:
        data = get_plan_monthly(
            country=country,
            city=city,
            lob_base=lob_base,
            segment=segment,
            year=year
        )
        return {
            "data": data,
            "total_periods": len(data),
            "year": year
        }
    except Exception as e:
        logger.error(f"Error al obtener Plan monthly: {e}")
        raise HTTPException(status_code=500, detail=f"Error al obtener Plan monthly: {str(e)}")

@router.get("/compare/overlap-monthly")
async def get_overlap_monthly_endpoint(
    country: Optional[str] = Query(None, description="Filtrar por país"),
    city: Optional[str] = Query(None, description="Filtrar por ciudad"),
    lob_base: Optional[str] = Query(None, description="Filtrar por línea de negocio base"),
    segment: Optional[str] = Query(None, description="Filtrar por segmento (b2b, b2c)"),
    year: Optional[int] = Query(None, description="Año específico para overlap (opcional)")
):
    """
    Obtiene comparación Plan vs Real SOLO para meses donde hay overlap temporal.
    Retorna lista vacía si no hay overlap, sin error.
    """
    try:
        data = get_overlap_monthly(
            country=country,
            city=city,
            lob_base=lob_base,
            segment=segment,
            year=year
        )
        return {
            "data": data,
            "total_periods": len(data),
            "has_overlap": len(data) > 0
        }
    except Exception as e:
        logger.error(f"Error al obtener overlap monthly: {e}")
        # Tolerante: retornar lista vacía en caso de error
        return {
            "data": [],
            "total_periods": 0,
            "has_overlap": False,
            "error": str(e)
        }


# --- BUSINESS_SLICE (capa ejecutiva; REAL-only) --------------------------------


@router.get("/business-slice/filters")
async def business_slice_filters(request: Request):
    try:
        return await _run_sync_request(request, get_business_slice_filters)
    except Exception as e:
        logger.error("business-slice/filters: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/business-slice/monthly")
async def business_slice_monthly(
    request: Request,
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    business_slice: Optional[str] = Query(None, description="Nombre de tajada (business_slice_name)"),
    fleet: Optional[str] = Query(None, description="fleet_display_name"),
    subfleet: Optional[str] = Query(None, description="subfleet_name"),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None, ge=1, le=12),
    limit: int = Query(2000, ge=1, le=10000),
):
    try:
        def _monthly_bundle():
            rows = get_business_slice_monthly(
                country=country,
                city=city,
                business_slice=business_slice,
                fleet=fleet,
                subfleet=subfleet,
                year=year,
                month=month,
                limit=limit,
            )
            rows = append_unmapped_bucket_rows(
                rows,
                "monthly",
                country=country,
                city=city,
                year=year,
                month=month,
                business_slice=business_slice,
                fleet=fleet,
                subfleet=subfleet,
            )
            meta = enrich_business_slice_matrix_meta(
                get_business_slice_matrix_freshness_meta(),
                "monthly",
                rows,
                country=country,
                city=city,
                business_slice=business_slice,
                fleet=fleet,
                subfleet=subfleet,
                year=year,
                month=month,
                fact_layer={
                    "grain": "monthly",
                    "status": "ok",
                    "source": "month_fact",
                    "source_table": "ops.real_business_slice_month_fact",
                },
            )
            return rows, meta

        data, meta = await _run_sync_request(request, _monthly_bundle)
        return {"data": data, "total": len(data), "meta": meta}
    except Exception as e:
        logger.error("business-slice/monthly: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/business-slice/coverage")
async def business_slice_coverage(
    request: Request,
    year: Optional[int] = Query(None),
    limit_months: int = Query(36, ge=1, le=120),
):
    try:
        return await _run_sync_request(
            request, get_business_slice_coverage, year=year, limit_months=limit_months
        )
    except Exception as e:
        logger.error("business-slice/coverage: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/business-slice/coverage-summary")
async def business_slice_coverage_summary(
    request: Request,
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None, ge=1, le=12),
):
    try:
        return await _run_sync_request(
            request,
            get_business_slice_coverage_summary,
            country=country, city=city, year=year, month=month,
        )
    except Exception as e:
        msg = str(e).lower()
        logger.error("business-slice/coverage-summary: %s", e)
        if "timeout" in msg or "canceling statement" in msg or "disk full" in msg or "no space" in msg:
            raise HTTPException(
                status_code=503,
                detail="coverage-summary: base de datos ocupada, timeout o recurso insuficiente; reintente más tarde.",
            ) from e
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/business-slice/unmatched")
async def business_slice_unmatched(
    request: Request,
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    limit: int = Query(80, ge=1, le=500),
):
    try:
        data = await _run_sync_request(
            request, get_business_slice_unmatched, country=country, city=city, limit=limit
        )
        return {"data": data, "total": len(data)}
    except Exception as e:
        logger.error("business-slice/unmatched: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/business-slice/conflicts")
async def business_slice_conflicts(
    request: Request,
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    limit: int = Query(80, ge=1, le=500),
):
    try:
        data = await _run_sync_request(
            request, get_business_slice_conflicts, country=country, city=city, limit=limit
        )
        return {"data": data, "total": len(data)}
    except Exception as e:
        logger.error("business-slice/conflicts: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/business-slice/subfleets")
async def business_slice_subfleets(request: Request):
    try:
        data = await _run_sync_request(request, get_business_slice_subfleets)
        return {"data": data, "total": len(data)}
    except Exception as e:
        logger.error("business-slice/subfleets: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/business-slice/weekly")
async def business_slice_weekly(
    request: Request,
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    business_slice: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    limit: int = Query(1500, ge=1, le=10000),
):
    try:
        def _weekly_bundle():
            rows, fact_layer = get_business_slice_weekly(
                country=country,
                city=city,
                business_slice=business_slice,
                year=year,
                limit=limit,
            )
            if fact_layer.get("status") == "ok":
                rows = append_unmapped_bucket_rows(
                    rows,
                    "weekly",
                    country=country,
                    city=city,
                    year=year,
                    month=None,
                    business_slice=business_slice,
                    fleet=None,
                    subfleet=None,
                )
            meta = enrich_business_slice_matrix_meta(
                get_business_slice_matrix_freshness_meta(),
                "weekly",
                rows,
                country=country,
                city=city,
                business_slice=business_slice,
                fleet=None,
                subfleet=None,
                year=year,
                month=None,
                fact_layer=fact_layer,
            )
            return rows, meta

        data, meta = await _run_sync_request(request, _weekly_bundle)
        return {"data": data, "total": len(data), "meta": meta}
    except Exception as e:
        logger.error("business-slice/weekly: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/business-slice/daily")
async def business_slice_daily(
    request: Request,
    country: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    business_slice: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None, ge=1, le=12),
    limit: int = Query(2000, ge=1, le=10000),
):
    try:
        def _daily_bundle():
            rows, fact_layer = get_business_slice_daily(
                country=country,
                city=city,
                business_slice=business_slice,
                year=year,
                month=month,
                limit=limit,
            )
            if fact_layer.get("status") == "ok":
                rows = append_unmapped_bucket_rows(
                    rows,
                    "daily",
                    country=country,
                    city=city,
                    year=year,
                    month=month,
                    business_slice=business_slice,
                    fleet=None,
                    subfleet=None,
                )
            meta = enrich_business_slice_matrix_meta(
                get_business_slice_matrix_freshness_meta(),
                "daily",
                rows,
                country=country,
                city=city,
                business_slice=business_slice,
                fleet=None,
                subfleet=None,
                year=year,
                month=month,
                fact_layer=fact_layer,
            )
            return rows, meta

        data, meta = await _run_sync_request(request, _daily_bundle)
        return {"data": data, "total": len(data), "meta": meta}
    except Exception as e:
        logger.error("business-slice/daily: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/business-slice/plan-join-stub")
async def business_slice_plan_join_stub(
    request: Request, limit: int = Query(500, ge=1, le=2000)
):
    """Contrato futuro Plan vs Real por business_slice (sin datos Plan)."""
    try:
        data = await _run_sync_request(request, get_plan_business_slice_stub, limit=limit)
        return {"data": data, "total": len(data)}
    except Exception as e:
        logger.error("business-slice/plan-join-stub: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/business-slice/omniview")
async def business_slice_omniview(
    request: Request,
    granularity: Literal["monthly", "weekly", "daily"] = Query(
        ...,
        description="Granularidad temporal: monthly (MoM), weekly (WoW), daily (vs mismo día -7).",
    ),
    period: Optional[str] = Query(
        None,
        description="Ancla: YYYY-MM (mensual/semanal) o YYYY-MM-DD (diario obligatorio). Opcional = defecto según granularidad.",
    ),
    country: Optional[str] = Query(None, description="Obligatorio para weekly y daily."),
    city: Optional[str] = Query(None),
    business_slice: Optional[str] = Query(None, description="Filtro business_slice_name"),
    fleet: Optional[str] = Query(None, description="Filtro fleet_display_name"),
    subfleet: Optional[str] = Query(None, description="Filtro subfleet_name"),
    include_subfleets: bool = Query(False, description="Si false, excluye filas is_subfleet"),
    daily_window_days: int = Query(
        90,
        ge=1,
        le=120,
        description="Tope validado (1–120); reservado para extensiones; V1 diario compara un solo día.",
    ),
    limit_rows: int = Query(2000, ge=1, le=10000),
    include_previous_only_rows: bool = Query(
        False,
        description="Incluye dimensiones solo presentes en el periodo anterior (current vacío).",
    ),
):
    """
    Omniview Business Slice (REAL): periodo actual vs anterior, métricas V1, deltas y señales.

    No reemplaza `/business-slice/monthly|weekly|daily`; agrega comparativo y rollups listos para UI.
    Weekly/daily están acotados: `country` obligatorio (guardrail de performance sobre la vista resuelta).
    Mensual: detalle desde `ops.real_business_slice_month_fact`; subtotales/totales desde vista resuelta.
    """
    try:
        return await _run_sync_request(
            request,
            get_business_slice_omniview,
            granularity=granularity,
            period=period,
            country=country,
            city=city,
            business_slice=business_slice,
            fleet=fleet,
            subfleet=subfleet,
            include_subfleets=include_subfleets,
            daily_window_days=daily_window_days,
            limit_rows=limit_rows,
            include_previous_only_rows=include_previous_only_rows,
        )
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        logger.exception("business-slice/omniview")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/business-slice/matrix-operational-trust")
async def business_slice_matrix_operational_trust(request: Request):
    """Trust operativo OK|warning|blocked para Omniview Matrix (validación de integridad)."""
    try:
        return await _run_sync_request(request, get_matrix_operational_trust_api_payload)
    except Exception as e:
        logger.exception("business-slice/matrix-operational-trust")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/business-slice/backfill-progress")
async def business_slice_backfill_progress():
    """Estado en tiempo real del backfill (mes actual, chunk actual, contadores)."""
    from app.services.backfill_runner import get_progress
    return get_progress()


@router.post("/business-slice/backfill")
async def business_slice_backfill_start(payload: dict = Body(...)):
    """Dispara un backfill de day_fact + week_fact para un rango de meses.
    Body: { from_date: "2025-01", to_date: "2025-06", with_week: true }
    """
    from app.services.backfill_runner import start_backfill, is_running
    from datetime import date as date_type

    if is_running():
        raise HTTPException(status_code=409, detail="Ya hay un backfill corriendo. Esperá a que termine.")

    def _parse_ym(s: str) -> date_type:
        s = str(s).strip()
        if len(s) == 7 and s[4] == "-":
            y, m = int(s[:4]), int(s[5:7])
            return date_type(y, m, 1)
        raise ValueError(f"Formato inválido: {s!r}. Usar YYYY-MM")

    try:
        from_date = _parse_ym(payload.get("from_date", ""))
        to_date   = _parse_ym(payload.get("to_date", ""))
    except (ValueError, TypeError) as e:
        raise HTTPException(status_code=422, detail=str(e))

    with_week = bool(payload.get("with_week", True))
    started = start_backfill(from_date, to_date, with_week=with_week)
    if not started:
        raise HTTPException(status_code=409, detail="No se pudo iniciar el backfill.")
    return {"ok": True, "from_date": str(from_date)[:7], "to_date": str(to_date)[:7], "with_week": with_week}


@router.post("/business-slice/backfill-cancel")
async def business_slice_backfill_cancel():
    """Cancela el backfill en curso (espera a que termine el chunk actual)."""
    from app.services.backfill_runner import cancel, is_running
    if not is_running():
        return {"ok": False, "detail": "No hay backfill corriendo"}
    cancel()
    return {"ok": True, "detail": "Cancelación solicitada — termina el chunk actual y para"}


@router.get("/business-slice/fact-status")
async def business_slice_fact_status():
    """Estado actual de las 3 FACT tables (month/week/day): qué meses están cargados y cuántas filas."""
    import psycopg2
    from psycopg2.extras import RealDictCursor
    from app.db.connection import get_db
    try:
        def _query():
            with get_db() as conn:
                cur = conn.cursor(cursor_factory=RealDictCursor)
                cur.execute("""
                    WITH month_status AS (
                        SELECT
                            month::text AS period,
                            SUM(trips_completed)::bigint AS trips,
                            COUNT(*)::int AS slices,
                            MAX(loaded_at) AS loaded_at
                        FROM ops.real_business_slice_month_fact
                        GROUP BY month
                        ORDER BY month
                    ),
                    day_months AS (
                        SELECT
                            date_trunc('month', trip_date)::date::text AS period,
                            SUM(trips_completed)::bigint AS trips,
                            COUNT(DISTINCT trip_date)::int AS days,
                            MAX(loaded_at) AS loaded_at
                        FROM ops.real_business_slice_day_fact
                        GROUP BY 1
                        ORDER BY 1
                    ),
                    week_months AS (
                        SELECT
                            date_trunc('month', week_start)::date::text AS period,
                            SUM(trips_completed)::bigint AS trips,
                            COUNT(DISTINCT week_start)::int AS weeks,
                            MAX(loaded_at) AS loaded_at
                        FROM ops.real_business_slice_week_fact
                        GROUP BY 1
                        ORDER BY 1
                    )
                    SELECT
                        ms.period,
                        ms.trips   AS month_trips,
                        ms.slices  AS month_slices,
                        ms.loaded_at AS month_loaded_at,
                        dm.trips   AS day_trips,
                        dm.days    AS day_days,
                        dm.loaded_at AS day_loaded_at,
                        wm.trips   AS week_trips,
                        wm.weeks   AS week_weeks,
                        wm.loaded_at AS week_loaded_at
                    FROM month_status ms
                    LEFT JOIN day_months dm USING (period)
                    LEFT JOIN week_months wm USING (period)
                    ORDER BY ms.period
                """)
                rows = [dict(r) for r in cur.fetchall()]
                for r in rows:
                    for k in ("month_loaded_at", "day_loaded_at", "week_loaded_at"):
                        if r.get(k) is not None:
                            r[k] = r[k].isoformat()
                # Totales globales
                cur.execute("SELECT COUNT(DISTINCT month) FROM ops.real_business_slice_month_fact")
                total_months = (cur.fetchone() or {}).get("count", 0)
                cur.execute("SELECT COUNT(DISTINCT date_trunc('month', trip_date)::date) FROM ops.real_business_slice_day_fact")
                total_day_months = (cur.fetchone() or {}).get("count", 0)
                cur.execute("SELECT COUNT(DISTINCT date_trunc('month', week_start)::date) FROM ops.real_business_slice_week_fact")
                total_week_months = (cur.fetchone() or {}).get("count", 0)
                cur.close()
            return {
                "months": rows,
                "summary": {
                    "total_months_in_month_fact": int(total_months or 0),
                    "total_months_in_day_fact": int(total_day_months or 0),
                    "total_months_in_week_fact": int(total_week_months or 0),
                    "complete": int(total_day_months or 0) >= int(total_months or 1),
                }
            }
        return await _run_sync(_query)
    except Exception as e:
        logger.exception("business-slice/fact-status")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/business-slice/matrix-issue-action")
async def business_slice_matrix_issue_action(payload: dict = Body(...)):
    """Registra ejecución o resolución de una acción sobre un issue de Omniview Matrix."""
    try:
        issue = payload.get("issue") if isinstance(payload.get("issue"), dict) else {}
        action_status = str(payload.get("action_status") or "").strip().lower()
        notes = payload.get("notes")
        if not issue:
            raise HTTPException(status_code=422, detail="issue es obligatorio")
        if action_status not in ("executed", "resolved"):
            raise HTTPException(status_code=422, detail="action_status debe ser executed o resolved")
        row = await _run_sync(log_omniview_issue_action, issue, action_status, notes)
        return {"status": "ok", "action": row}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("business-slice/matrix-issue-action")
        raise HTTPException(status_code=500, detail=str(e))


# --- Revenue Proxy Coverage ---

@router.get("/revenue-proxy/coverage")
async def revenue_proxy_coverage():
    """Cobertura de revenue real vs proxy por mes/país/ciudad.
    Lee de ops.v_real_revenue_proxy_coverage (vista creada en migración 120).
    """
    from app.db.connection import get_db
    from psycopg2.extras import RealDictCursor
    try:
        def _fetch():
            with get_db() as conn:
                cur = conn.cursor(cursor_factory=RealDictCursor)
                cur.execute("""
                    SELECT * FROM ops.v_real_revenue_proxy_coverage
                    ORDER BY trip_month DESC, country, city
                    LIMIT 200
                """)
                rows = [dict(r) for r in cur.fetchall()]
                cur.close()
                return rows
        data = await _run_sync(_fetch)
        return {"data": data, "total": len(data)}
    except Exception as e:
        logger.error("revenue-proxy/coverage: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/revenue-proxy/config")
async def revenue_proxy_config():
    """Configuración actual de comisión proxy (ops.yego_commission_proxy_config)."""
    from app.db.connection import get_db
    from psycopg2.extras import RealDictCursor
    try:
        def _fetch():
            with get_db() as conn:
                cur = conn.cursor(cursor_factory=RealDictCursor)
                cur.execute("""
                    SELECT id, country, city, park_id, tipo_servicio,
                           commission_pct, valid_from, valid_to,
                           priority, is_active, notes
                    FROM ops.yego_commission_proxy_config
                    WHERE is_active
                    ORDER BY priority DESC, valid_from DESC
                """)
                rows = [dict(r) for r in cur.fetchall()]
                cur.close()
                return rows
        data = await _run_sync(_fetch)
        return {"data": data, "total": len(data)}
    except Exception as e:
        logger.error("revenue-proxy/config: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# --- Revenue Quality / Hardening ---

@router.get("/revenue-quality/check")
async def revenue_quality_check():
    """Ejecuta validación completa de calidad de revenue y retorna alertas."""
    from app.services.revenue_quality_service import run_revenue_quality_check, persist_alerts
    try:
        result = await _run_sync(run_revenue_quality_check)
        await _run_sync(persist_alerts, result.get("alerts", []))
        return result
    except Exception as e:
        logger.error("revenue-quality/check: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/revenue-quality/alerts")
async def revenue_quality_alerts(limit: int = Query(50, ge=1, le=200)):
    """Alertas recientes de calidad de revenue (persistidas)."""
    from app.services.revenue_quality_service import get_latest_alerts
    try:
        data = await _run_sync(get_latest_alerts, limit)
        return {"data": data, "total": len(data)}
    except Exception as e:
        logger.error("revenue-quality/alerts: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/revenue-quality/by-city")
async def revenue_quality_by_city(days: int = Query(7, ge=1, le=90)):
    """Resumen de calidad de revenue por ciudad (últimos N días)."""
    from app.services.revenue_quality_service import get_revenue_quality_by_city
    try:
        data = await _run_sync(get_revenue_quality_by_city, days)
        return {"data": data, "total": len(data)}
    except Exception as e:
        logger.error("revenue-quality/by-city: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# --- Action Engine ---

@router.get("/action-engine/run")
async def action_engine_run():
    """Ejecuta el Action Engine: genera y persiste acciones priorizadas del día."""
    from app.services.action_engine_service import run_action_engine, persist_action_output
    try:
        result = await _run_sync(run_action_engine)
        persisted = await _run_sync(persist_action_output, result)
        result["persisted"] = persisted
        return result
    except Exception as e:
        logger.error("action-engine/run: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/action-engine/today")
async def action_engine_today(
    city: Optional[str] = Query(None),
    severity: Optional[str] = Query(None),
    limit: int = Query(50, ge=1, le=200),
):
    """Acciones del día ordenadas por prioridad."""
    from app.services.action_engine_service import get_today_actions
    try:
        data = await _run_sync(get_today_actions, city, severity, limit)
        return {"date": str(__import__('datetime').date.today()), "actions": data, "total": len(data)}
    except Exception as e:
        logger.error("action-engine/today: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/action-engine/catalog")
async def action_engine_catalog():
    """Catálogo de acciones operativas disponibles."""
    from app.db.connection import get_db
    from psycopg2.extras import RealDictCursor
    try:
        def _fetch():
            with get_db() as conn:
                cur = conn.cursor(cursor_factory=RealDictCursor)
                cur.execute("""
                    SELECT action_id, action_name, action_type, description,
                           trigger_metric, trigger_condition, severity,
                           suggested_owner, suggested_channel, expected_impact, is_active
                    FROM ops.action_catalog
                    WHERE is_active
                    ORDER BY severity DESC, action_name
                """)
                rows = [dict(r) for r in cur.fetchall()]
                cur.close()
                return rows
        data = await _run_sync(_fetch)
        return {"data": data, "total": len(data)}
    except Exception as e:
        logger.error("action-engine/catalog: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/action-engine/log")
async def action_engine_log_execution(
    action_output_id: int = Query(...),
    action_id: str = Query(...),
    owner: str = Query(...),
    status: str = Query("pending"),
    notes: Optional[str] = Query(None),
):
    """Registra tracking de ejecución de una acción."""
    from app.services.action_engine_service import log_action_execution
    try:
        new_id = await _run_sync(
            log_action_execution, action_output_id, action_id, owner, status, notes
        )
        return {"id": new_id, "status": "logged"}
    except Exception as e:
        logger.error("action-engine/log: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


def _action_plan_item(row: dict) -> dict:
    """Shape estable para UI / integraciones (fase 8)."""
    return {
        "id": row.get("id"),
        "country": row.get("country"),
        "city": row.get("city") or "",
        "park_id": row.get("park_id"),
        "action_id": row.get("action_id"),
        "action": row.get("action_name"),
        "volume": row.get("suggested_volume"),
        "target_segment": row.get("target_segment"),
        "priority_score": float(row["priority_score"])
        if row.get("priority_score") is not None
        else None,
        "severity": row.get("severity"),
        "steps": row.get("suggested_playbook_text"),
        "expected_impact": row.get("expected_impact"),
        "status": row.get("status"),
        "playbook_id": row.get("suggested_playbook_id"),
    }


@router.get("/action-plan/today")
async def action_plan_today(
    limit: int = Query(100, ge=1, le=500),
):
    """Plan operativo del día (prioridad DESC). Requiere corrida previa del orquestador."""
    from datetime import date as date_cls
    from app.services.action_orchestrator_service import get_action_plans

    try:
        rows = await _run_sync(get_action_plans, date_cls.today(), None, limit, 0)
        d = str(date_cls.today())
        return {
            "date": d,
            "plans": [_action_plan_item(dict(r)) for r in rows],
            "total": len(rows),
        }
    except Exception as e:
        logger.error("action-plan/today: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/action-plan/top")
async def action_plan_top(
    limit: int = Query(20, ge=1, le=200),
    plan_date: Optional[str] = Query(None, description="YYYY-MM-DD; default hoy"),
):
    """Top del plan por priority_score (mismo día que today salvo plan_date)."""
    from datetime import date as date_cls
    from app.services.action_orchestrator_service import get_action_plans

    try:
        pd = date_cls.today()
        if plan_date:
            pd = date_cls.fromisoformat(plan_date)
        rows = await _run_sync(get_action_plans, pd, None, limit, 0)
        return {
            "date": str(pd),
            "plans": [_action_plan_item(dict(r)) for r in rows],
            "total": len(rows),
        }
    except ValueError as e:
        raise HTTPException(status_code=422, detail=f"plan_date inválida: {e}")
    except Exception as e:
        logger.error("action-plan/top: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/action-plan")
async def action_plan_filter(
    city: Optional[str] = Query(None),
    plan_date: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0, le=10_000),
):
    """Plan filtrado por ciudad y fecha."""
    from datetime import date as date_cls
    from app.services.action_orchestrator_service import get_action_plans

    try:
        pd = date_cls.today()
        if plan_date:
            pd = date_cls.fromisoformat(plan_date)
        rows = await _run_sync(get_action_plans, pd, city, limit, offset)
        return {
            "date": str(pd),
            "city_filter": city,
            "plans": [_action_plan_item(dict(r)) for r in rows],
            "total": len(rows),
        }
    except ValueError as e:
        raise HTTPException(status_code=422, detail=f"plan_date inválida: {e}")
    except Exception as e:
        logger.error("action-plan: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/action-plan/run")
async def action_plan_run():
    """Regenera ops.action_plan_daily para hoy a partir de ops.action_engine_output."""
    from app.services.action_orchestrator_service import run_action_orchestrator

    try:
        result = await _run_sync(run_action_orchestrator)
        return {"ok": True, **result}
    except Exception as e:
        logger.error("action-plan/run: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/action-plan/log")
async def action_plan_log_execution(
    action_plan_id: int = Query(...),
    action_id: str = Query(...),
    owner: str = Query(...),
    status: str = Query("pending"),
    notes: Optional[str] = Query(None),
):
    """Tracking humano sobre una fila del plan (no dispara acciones externas)."""
    from app.services.action_orchestrator_service import log_plan_execution

    try:
        new_id = await _run_sync(
            log_plan_execution, action_plan_id, action_id, owner, status, notes
        )
        return {"id": new_id, "status": "logged", "target": "action_plan_daily"}
    except Exception as e:
        logger.error("action-plan/log: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# --- Learning Engine ---

@router.get("/action-learning/effectiveness")
async def action_learning_effectiveness(
    action_id: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    country: Optional[str] = Query(None),
    limit: int = Query(200, ge=1, le=1000),
):
    """Efectividad histórica por acción / ciudad / país."""
    from app.services.action_learning_service import get_effectiveness

    try:
        rows = await _run_sync(get_effectiveness, action_id, city, country, limit)
        return {"data": rows, "total": len(rows)}
    except Exception as e:
        logger.error("action-learning/effectiveness: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/action-learning/executions")
async def action_learning_executions(
    action_id: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0, le=10_000),
):
    """Acciones evaluadas con before/after/delta."""
    from app.services.action_learning_service import get_evaluated_executions

    try:
        rows = await _run_sync(get_evaluated_executions, action_id, city, limit, offset)
        return {"data": rows, "total": len(rows)}
    except Exception as e:
        logger.error("action-learning/executions: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/action-learning/evaluation-rules")
async def action_learning_evaluation_rules(
    active_only: bool = Query(True),
):
    """Reglas de evaluación por acción (métrica, dirección, ventana, umbral)."""
    from app.services.action_learning_service import get_evaluation_rules

    try:
        rows = await _run_sync(get_evaluation_rules, active_only)
        return {"data": rows, "total": len(rows)}
    except Exception as e:
        logger.error("action-learning/evaluation-rules: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/action-learning/evaluate")
async def action_learning_evaluate(
    force: bool = Query(False, description="Re-evaluar acciones ya evaluadas"),
    limit: int = Query(500, ge=1, le=5000),
):
    """Ejecuta evaluación de acciones completadas (escribe result en execution_log)."""
    from app.services.action_learning_service import evaluate_executions

    try:
        result = await _run_sync(evaluate_executions, "done", force, limit)
        return {"ok": True, **result}
    except Exception as e:
        logger.error("action-learning/evaluate: %s", e)
        raise HTTPException(status_code=500, detail=str(e))
